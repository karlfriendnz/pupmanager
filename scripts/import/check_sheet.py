"""
plan.json → a THREE-TAB check sheet: Clients, Packages, Classes.

This is the sheet a human actually checks before an import goes live. It is
deliberately not the same artefact as `review.py`'s eight-tab workbook: that one
exists to explain how every decision was reached, this one exists to be read
down a column and signed off.

Rows that need a human eye are shaded, and every one of them says why in an
Issue column — a colour with no explanation just moves the question along.

    red    something is probably wrong and would import badly
    amber  worth a look; the import will proceed either way
    none   no known problem

Opens in Excel, Numbers or Google Sheets (upload it — Sheets converts .xlsx on
open, keeping the shading and the filters).

Usage:
    python3 scripts/import/check_sheet.py <plan.json> <out.xlsx>
"""

import json
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Shading. Muted on purpose — a sheet where everything shouts is a sheet nobody
# reads to the bottom.
RED = PatternFill("solid", fgColor="F8D7DA")
AMBER = PatternFill("solid", fgColor="FFF3CD")
HEADER = PatternFill("solid", fgColor="1F2933")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)

# Review kinds that mean "this will import wrongly unless someone looks".
# Everything else is amber.
BLOCKING = {
    'Two professionals share one email',
    'Person with no name',
    'Spreadsheet row could not be read',
    'Invented class may already exist',
    'Contact with no email — possible duplicate',
}


def issues_by_person(plan: dict) -> dict:
    """Map person id → list of issue strings, from their flags and needsReview."""
    out: dict[str, list[str]] = defaultdict(list)

    for p in plan['people']:
        for f in p.get('flags') or []:
            out[p['id']].append(f)

    # needsReview identifies people by display name, so build a name index. A
    # name can be ambiguous; attaching the note to every match is the safe
    # direction — better two people see a question than nobody does.
    by_name: dict[str, list[str]] = defaultdict(list)
    for p in plan['people']:
        name = (p.get('fullName') or '').strip().lower()
        if name:
            by_name[name].append(p['id'])

    for row in plan.get('needsReview') or []:
        who = (row.get('who') or '').strip().lower()
        note = f"{row.get('kind')}: {row.get('detail')}".strip()
        for pid in by_name.get(who, []):
            out[pid].append(note)

    return out


def severity(issues: list[str]) -> PatternFill | None:
    if not issues:
        return None
    return RED if any(any(b in i for b in BLOCKING) for i in issues) else AMBER


def write_sheet(ws, headers: list[str], rows: list[tuple[list, list[str]]], widths: list[int]):
    """rows = [(cells, issues)] — issues drive the shading and the last column."""
    ws.append(headers)
    for i, _ in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i)
        c.fill = HEADER
        c.font = HEADER_FONT
        c.alignment = Alignment(vertical='center')

    for cells, issues in rows:
        ws.append([*cells, ' · '.join(issues)])
        fill = severity(issues)
        if fill:
            for i in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=i).fill = fill

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{max(ws.max_row, 1)}'
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 22


def clients_tab(wb, plan, issues):
    ws = wb.create_sheet('Clients')
    headers = ['Name', 'Email', 'Phone', 'Address', 'Dogs', 'Classes', '1:1s',
               'Volunteer', 'Professional', 'Previous trainer', 'Business',
               'Credit', 'Status', 'Issue']
    rows = []
    for p in plan['people']:
        cf = p.get('customFields') or {}
        dogs = ', '.join(
            f"{d.get('name') or '(unnamed)'}" + (f" ({d['breed']})" if d.get('breed') else '')
            for d in p.get('dogs') or []
        )
        enrol = len(p.get('enrolments') or [])
        consults = len(p.get('oneToOneConsults') or [])
        iss = list(issues.get(p['id'], []))
        if p.get('emailIsPlaceholder'):
            iss.append('No email — placeholder generated, will not merge')
        rows.append((
            [p.get('fullName') or '', '' if p.get('emailIsPlaceholder') else (p.get('email') or ''),
             p.get('phone') or '', p.get('address') or '', dogs, enrol, consults,
             cf.get('Volunteer', ''), cf.get('Professional', ''), cf.get('Previous trainer', ''),
             cf.get('Business', '') or p.get('company') or '', cf.get('Credit', ''),
             'Active' if (enrol or consults) else 'New'],
            iss,
        ))
    write_sheet(ws, headers, rows, [26, 30, 16, 34, 30, 8, 6, 10, 12, 15, 22, 22, 9, 60])
    return len(rows), sum(1 for _, i in rows if i)


def packages_tab(wb, plan, issues):
    """One row per 1:1 consult — these become ClientPackages on import."""
    ws = wb.create_sheet('Packages')
    headers = ['Client', 'Email', 'Service', 'Code', 'Month', 'Year', 'From the file', 'Issue']
    rows = []
    for p in plan['people']:
        for c in p.get('oneToOneConsults') or []:
            iss = []
            if not c.get('month') or not c.get('year'):
                iss.append('No date in the source — will import with a placeholder date')
            rows.append((
                [p.get('fullName') or '', '' if p.get('emailIsPlaceholder') else (p.get('email') or ''),
                 c.get('service') or '', c.get('code') or '', c.get('month') or '',
                 c.get('year') or '', c.get('raw') or ''],
                iss,
            ))
    rows.sort(key=lambda r: (r[0][2], r[0][0]))
    write_sheet(ws, headers, rows, [26, 30, 26, 12, 12, 8, 44, 52])
    return len(rows), sum(1 for _, i in rows if i)


def classes_tab(wb, plan, issues):
    ws = wb.create_sheet('Classes')
    headers = ['Class', 'Level', 'Code', 'Where it came from', 'Day', 'Date',
               'Time', 'Location', 'Enrolments', 'Issue']
    rows = []
    review_by_who = defaultdict(list)
    for row in plan.get('needsReview') or []:
        review_by_who[(row.get('who') or '').strip().lower()].append(
            f"{row.get('kind')}: {row.get('detail')}"
        )

    for c in plan['courses']:
        iss = list(c.get('flags') or [])
        iss += review_by_who.get((c.get('name') or '').strip().lower(), [])
        origin = 'From the spreadsheet' if c.get('origin') == 'spreadsheet' else 'Reconstructed from a note'
        if c.get('origin') != 'spreadsheet':
            iss.append('Invented from a contact note — check it is not a duplicate, then rename')
        rows.append((
            [c.get('name') or '', c.get('level') or '', c.get('code') or '', origin,
             c.get('day') or '', c.get('date') or '', c.get('time') or '',
             c.get('location') or '', c.get('enrolmentCount') or 0],
            iss,
        ))
    rows.sort(key=lambda r: (r[0][3], r[0][1], r[0][0]))
    write_sheet(ws, headers, rows, [40, 18, 8, 26, 12, 16, 18, 16, 11, 64])
    return len(rows), sum(1 for _, i in rows if i)


def main(plan_path: Path, out_path: Path) -> None:
    plan = json.loads(plan_path.read_text(encoding='utf-8'))
    issues = issues_by_person(plan)

    wb = Workbook()
    wb.remove(wb.active)
    counts = {
        'Clients': clients_tab(wb, plan, issues),
        'Packages': packages_tab(wb, plan, issues),
        'Classes': classes_tab(wb, plan, issues),
    }
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    print(f'Wrote {out_path}')
    for tab, (total, flagged) in counts.items():
        print(f'  {tab:<9} {total:>5} rows   {flagged:>4} need a look')


if __name__ == '__main__':
    if len(sys.argv) != 3:
        sys.exit(__doc__)
    main(Path(sys.argv[1]), Path(sys.argv[2]))
