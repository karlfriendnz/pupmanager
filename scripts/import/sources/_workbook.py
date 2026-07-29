"""
Reading a course-intake WORKBOOK into flat records.

WHY THIS IS SEPARATE FROM THE SOURCE MODULE
-------------------------------------------
Two different jobs live in one file otherwise. This one is "what does this
workbook literally say?" — find the course header rows, work out which rows sit
under each of them, classify each row's cells, and deduplicate the people. The
source module next door is "how do we express that as fragments the merge
engine can use?".

Keeping them apart means the awkward part — and it is all awkward: ALL-CAPS
category headers, note sentences that belong to the row above, a bare year
sitting alone in column A, a sheet whose column order changes half way down —
can be read, and fixed, without also holding the fragment protocol in your
head.

WHAT MAKES A COURSE HEADER
--------------------------
"STARTS ..." is unambiguous. Without that prefix we demand three things at
once: the line starts with a weekday, carries a time, AND carries a month.
Any two of those matches a sentence like "Last class to run after Xmas on
Saturday 4th January at 9am" — which is a note, and treating it as a header
opens a phantom course and reassigns every enrolment under it.
"""

from __future__ import annotations

import re

import openpyxl
from openpyxl.utils import get_column_letter

from lib import parsing

WEEKDAY = r"(monday|tuesday|wednesday|thursday|friday|saturday|sunday)"
MONTH = (r"(january|february|march|april|may|june|july|august|september|"
         r"october|november|december)")
TIME_RE = re.compile(r"\d{1,2}\s*[:.]?\d{0,2}\s*(am|pm)", re.I)
HEADER_RE = re.compile(r"^\s*(starts\s+)?" + WEEKDAY + r"\b.*?\d", re.I)

DEFAULT_MAX_COL = 13          # A..L, plus one for safety

# A single name cell holding two people ("Jo & Sam", "Ann/Bob"). Not a spelling
# variant — a household, which the client has to split by hand.
COUPLE_SEP_RE = re.compile(r"[A-Za-z]\s*(?:&|/|\+| and )\s*[A-Za-z]")

VENUE_HINTS = ("bethlehem", "matua", "bch", "hall")

def simple_key(v) -> str:
    """Lowercase + collapse whitespace. Punctuation is KEPT, unlike norm_name.

    Used for this source's internal dedup keys. An email is a machine
    identifier — folding out its dots would merge john.smith@ with johnsmith@,
    which are two different mailboxes.
    """
    if not v:
        return ""
    return re.sub(r"\s+", " ", str(v).strip().lower())


def digits_only(v) -> str:
    return re.sub(r"\D", "", str(v)) if v else ""


def is_header_label_row(vals: dict, subject_labels=("name", "nname")) -> bool:
    a = vals.get("A")
    if not isinstance(a, str) or a.strip().lower() not in subject_labels:
        return False
    joined = " ".join(str(x).lower() for x in vals.values())
    return "email" in joined or "phone" in joined or "puppy" in joined


def is_bare_year(vals: dict) -> bool:
    if len(vals) != 1 or "A" not in vals:
        return False
    a = vals["A"]
    if isinstance(a, (int, float)) and 2000 <= int(a) <= 2100:
        return True
    return isinstance(a, str) and bool(re.fullmatch(r"20\d\d", a.strip()))


def is_course_header(a) -> bool:
    """
    A course header, not a sentence that happens to name a day.

    "STARTS ..." is unambiguous. Without it we demand: starts with a weekday,
    carries a time, and carries a month. That triple is what stops "Last class
    to run on Saturday 4th January at 9am" opening a new course.
    """
    if not isinstance(a, str):
        return False
    s = a.strip()
    if s.lower().startswith("starts "):
        return True
    if HEADER_RE.match(s) and TIME_RE.search(s) and re.search(MONTH, s, re.I):
        first = s.split()[0].lower()
        return bool(re.fullmatch(WEEKDAY, first))
    return False


def parse_header(a: str) -> dict:
    """Day / date / time / nickname out of a header string."""
    s = a.strip()
    nick = None
    m = re.search(r"['\"‘’“”]([^'\"‘’“”]+)['\"‘’“”]", s)
    if m:
        nick = m.group(1).strip()
    else:
        m2 = re.search(r"-\s*(The [A-Z][\w' ]+)$", s)
        if m2:
            nick = m2.group(1).strip()

    md = re.search(WEEKDAY, s, re.I)
    day = md.group(1).capitalize() if md else None

    mdate = re.search(r"(\d{1,2}(?:st|nd|rd|th)?\s+" + MONTH + r")", s, re.I)
    date = mdate.group(1).strip() if mdate else None

    mt = re.search(r"(\d{1,2}[:.]?\d{0,2}\s*(?:am|pm)?\s*[-–]\s*"
                   r"\d{1,2}[:.]?\d{0,2}\s*(?:am|pm))", s, re.I)
    if mt:
        time = mt.group(1).strip()
    else:
        mt2 = TIME_RE.search(s)
        time = mt2.group(0).strip() if mt2 else None

    return {"day": day, "date": date, "time": time, "nickname": nick, "raw": s}


def header_location(gval, hints=VENUE_HINTS) -> str | None:
    """
    A venue name in a header's spare column — or a note that happens to sit
    there. Only a short string containing a known venue word counts.
    """
    if not isinstance(gval, str):
        return None
    low = gval.lower()
    if (any(h in low for h in hints) and len(gval) < 30
            and not any(w in low for w in ("cancel", "sent", "welcome", "itinerary"))):
        return gval.strip()
    return None


def row_values(ws, r: int, max_col: int) -> dict:
    vals = {}
    for c in range(1, max_col):
        v = ws.cell(row=r, column=c).value
        if v is not None and not (isinstance(v, str) and v.strip() == ""):
            vals[get_column_letter(c)] = v
    return vals


_CACHE: dict[tuple, dict] = {}


def read(path, options: dict) -> dict:
    """Parse the workbook once per (path, options) — vocabulary() reads it too."""
    key = (str(path), repr(sorted(options.items())))
    if key in _CACHE:
        return _CACHE[key]

    wb = openpyxl.load_workbook(path, data_only=True)
    class_sheets = options.get("classSheets") or [
        s for s in wb.sheetnames if s.lower() != "waitlist"]
    waitlist_sheet = options.get("waitlistSheet", "Waitlist")
    max_col = int(options.get("maxColumn", DEFAULT_MAX_COL))
    location_column = options.get("locationColumn", "G")
    location_sheets = set(options.get("locationSheets") or [])
    venue_hints = tuple(options.get("venueHints") or VENUE_HINTS)
    sheet_prefix = options.get("sheetPrefix") or {
        s: s[:2] for s in class_sheets}

    breed_re = parsing.build_breed_re(options.get("breedWords") or [])
    suburbs = {s.lower() for s in (options.get("suburbs") or [])}

    courses, enrolments, needs_review, waitlist = [], [], [], []
    course_id = 0

    for sheet in class_sheets:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        cur_course = None
        for r in range(1, ws.max_row + 1):
            vals = row_values(ws, r, max_col)
            if not vals:
                continue
            a = vals.get("A")

            if is_course_header(a):
                course_id += 1
                h = parse_header(a)
                loc = None
                if sheet in location_sheets:
                    loc = header_location(vals.get(location_column), venue_hints)
                cur_course = course_id
                courses.append({"id": course_id, "sheet": sheet,
                                "nickname": h["nickname"], "day": h["day"],
                                "date": h["date"], "time": h["time"],
                                "location": loc, "headerRaw": h["raw"]})
                continue

            if is_header_label_row(vals) or is_bare_year(vals):
                continue
            if cur_course is None:
                continue

            rec = parsing.classify_row(vals, breed_re=breed_re, suburbs=suburbs)
            src_code = f"{sheet_prefix.get(sheet, sheet[:2])}{r}"

            # No way to reach this person at all. Kept for review rather than
            # imported as a contactless record.
            if rec["email"] is None and rec["phone"] is None:
                needs_review.append({"sheet": sheet, "row": r,
                                     "reason": "no email+phone",
                                     "raw": parsing.compact_raw(vals)})
                continue

            enrolments.append({
                "courseId": cur_course, "name": rec["name"], "email": rec["email"],
                "phone": rec["phone"], "address": rec["address"], "dog": rec["dog"],
                "breed": rec["breed"], "age": rec["age"],
                "dateBooked": rec["dateBooked"], "notes": rec["notes"],
                "sourceSheet": sheet, "sourceRow": r, "sourceCode": src_code,
            })

    if waitlist_sheet in wb.sheetnames:
        waitlist = _read_waitlist(wb[waitlist_sheet], max_col, breed_re, suburbs)

    clients = _dedup_clients(enrolments)
    out = {"courses": courses, "enrolments": enrolments, "clients": clients,
           "waitlist": waitlist, "needsReview": needs_review}
    _CACHE[key] = out
    return out


def _read_waitlist(ws, max_col: int, breed_re, suburbs) -> list[dict]:
    """
    The waitlist sheet: ALL-CAPS category headers, then leads, then loose note
    sentences that belong to the lead above them.
    """
    out, category, last_lead = [], None, None
    for r in range(1, ws.max_row + 1):
        vals = row_values(ws, r, max_col)
        if not vals:
            continue
        a = vals.get("A")
        if is_header_label_row(vals):
            continue

        if isinstance(a, str):
            letters = [ch for ch in a if ch.isalpha()]
            upfrac = (sum(ch.isupper() for ch in letters) / len(letters)
                      if letters else 0)
            if len(vals) <= 2 and upfrac > 0.7 and len(a) < 40 and letters:
                category, last_lead = a.strip(), None
                continue
        if category is None:
            continue

        rec = parsing.classify_row(vals, breed_re=breed_re, suburbs=suburbs)
        aname = rec["name"] if isinstance(rec["name"], str) else ""
        low = aname.lower()
        is_note_sentence = (
            aname and rec["email"] is None and rec["phone"] is None
            and (any(t in low for t in ("hoping", "wants a", "have let",
                                        "let them", "let her", "let him"))
                 or len(aname.split()) > 4)
        )
        if is_note_sentence and last_lead is not None:
            extra = parsing.compact_raw(vals)
            last_lead["notes"] = ((last_lead["notes"] + " || " + extra)
                                  if last_lead["notes"] else extra)
            continue
        if not aname:
            continue

        lead = {"category": category, "name": aname, "phone": rec["phone"],
                "email": rec["email"], "address": rec["address"],
                "dog": rec["dog"], "breedAge": rec["breed"] or rec["age"],
                "notes": rec["notes"], "sourceRow": r}
        out.append(lead)
        last_lead = lead
    return out


def _dedup_clients(enrolments) -> list[dict]:
    """
    One person out of many enrolment rows.

    Email, then phone, then name — and name ONLY when there is neither email
    nor phone, because two family members legitimately share a surname but
    never a mailbox.
    """
    clients: list[dict] = []
    by_email: dict[str, dict] = {}
    by_phone: dict[str, dict] = {}
    by_name: dict[str, dict] = {}

    def get_client(rec):
        em = simple_key(rec["email"]) if rec["email"] else ""
        ph = digits_only(rec["phone"]) if rec["phone"] else ""
        nm = simple_key(rec["name"])
        if em and em in by_email:
            return by_email[em]
        if ph and ph in by_phone:
            return by_phone[ph]
        if not em and not ph and nm and nm in by_name:
            return by_name[nm]
        cl = {"name": rec["name"], "email": rec["email"], "phone": rec["phone"],
              "address": rec["address"], "dogs": [], "courseIds": [], "notes": "",
              "sourceRows": [], "multipleNamesOnKey": False,
              "_names": set(), "_notes": []}
        clients.append(cl)
        if em:
            by_email[em] = cl
        if ph:
            by_phone[ph] = cl
        if nm:
            by_name.setdefault(nm, cl)
        return cl

    for e in enrolments:
        cl = get_client(e)
        if not cl["email"] and e["email"]:
            cl["email"] = e["email"]
            by_email.setdefault(simple_key(e["email"]), cl)
        if not cl["phone"] and e["phone"]:
            cl["phone"] = e["phone"]
            by_phone.setdefault(digits_only(e["phone"]), cl)
        if not cl["address"] and e["address"]:
            cl["address"] = e["address"]
        if e["dog"] and e["dog"] not in cl["dogs"]:
            cl["dogs"].append(e["dog"])
        if e["courseId"] not in cl["courseIds"]:
            cl["courseIds"].append(e["courseId"])
        cl["sourceRows"].append(e["sourceCode"])
        cl["_names"].add(e["name"])
        if e["notes"]:
            cl["_notes"].append(f"[{e['sourceCode']}] {e['notes']}")
        if e["name"] and COUPLE_SEP_RE.search(str(e["name"])):
            cl["multipleNamesOnKey"] = True

    for cl in clients:
        if any(n and COUPLE_SEP_RE.search(str(n)) for n in cl["_names"]):
            cl["multipleNamesOnKey"] = True
        cl["notes"] = " || ".join(dict.fromkeys(cl["_notes"]))
        del cl["_names"]
        del cl["_notes"]
    return clients
