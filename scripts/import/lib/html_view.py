"""
Render an .xlsx workbook as ONE self-contained HTML page.

WHY THIS EXISTS
---------------
There is no spreadsheet application on the machine this runs on, so the review
workbook would otherwise be unopenable — and a review nobody can open does not
happen.

Deliberately local and offline. The content is real client names, home
addresses, phone numbers and, on the waitlist tab, notes about their children.
It must never be pasted into a hosted viewer to be read. One file, no network
requests, no fonts, no analytics: openable, mailable, and deletable.

A tab per sheet, sticky headers, and a per-tab filter box, because the way
this actually gets used is "find every row that mentions Bethlehem" and then
"find every row with no email".
"""

from __future__ import annotations

import html
from pathlib import Path

from openpyxl import load_workbook

CSS = """
:root { --bg:#fff; --fg:#1a1a1a; --muted:#666; --line:#e3e3e3; --accent:#2b6cb0;
        --head:#f7f8fa; --hit:#fffbe6; }
@media (prefers-color-scheme: dark) {
  :root { --bg:#15171a; --fg:#e8e8e8; --muted:#9aa0a6; --line:#2c3035;
          --accent:#7bb0e8; --head:#1d2025; --hit:#3a3520; }
}
* { box-sizing: border-box; }
body { margin:0; background:var(--bg); color:var(--fg); font:14px/1.45 -apple-system,
       BlinkMacSystemFont, "Segoe UI", Helvetica, Arial, sans-serif; }
header { padding:18px 20px 0; }
h1 { margin:0 0 2px; font-size:19px; }
.sub { color:var(--muted); font-size:13px; margin-bottom:14px; }
nav { display:flex; flex-wrap:wrap; gap:6px; padding:0 20px; border-bottom:1px solid var(--line); }
nav button { border:1px solid var(--line); border-bottom:none; background:transparent;
             color:var(--muted); padding:7px 13px; border-radius:7px 7px 0 0; cursor:pointer;
             font-size:13px; font-family:inherit; }
nav button:hover { color:var(--fg); }
nav button.on { background:var(--head); color:var(--fg); font-weight:600;
                box-shadow:inset 0 2px 0 var(--accent); }
nav button .n { color:var(--muted); font-weight:400; margin-left:5px; font-size:12px; }
.pane { display:none; padding:14px 20px 60px; }
.pane.on { display:block; }
.tools { display:flex; gap:10px; align-items:center; margin-bottom:10px; }
input[type=search] { flex:1; max-width:380px; padding:7px 10px; border:1px solid var(--line);
                     border-radius:7px; background:var(--bg); color:var(--fg);
                     font:inherit; font-size:13px; }
.count { color:var(--muted); font-size:12px; }
.scroll { overflow-x:auto; border:1px solid var(--line); border-radius:8px; }
table { border-collapse:collapse; width:100%; font-size:13px; }
th, td { text-align:left; padding:7px 10px; border-bottom:1px solid var(--line);
         vertical-align:top; white-space:pre-wrap; word-break:break-word; }
th { position:sticky; top:0; background:var(--head); font-weight:600; z-index:1;
     border-bottom:2px solid var(--line); }
tbody tr:hover { background:var(--hit); }
td:empty::after { content:"—"; color:var(--line); }
.empty { color:var(--muted); padding:20px; }
"""

JS = """
function show(i){
  document.querySelectorAll('nav button').forEach((b,n)=>b.classList.toggle('on',n===i));
  document.querySelectorAll('.pane').forEach((p,n)=>p.classList.toggle('on',n===i));
}
function filter(i){
  const pane = document.querySelectorAll('.pane')[i];
  const q = pane.querySelector('input').value.trim().toLowerCase();
  let shown = 0;
  pane.querySelectorAll('tbody tr').forEach(tr => {
    const hit = !q || tr.textContent.toLowerCase().includes(q);
    tr.style.display = hit ? '' : 'none';
    if (hit) shown++;
  });
  pane.querySelector('.count').textContent = shown + ' of ' + pane.dataset.total + ' rows';
}
document.addEventListener('keydown', e => {
  if (e.key === 'Escape') document.activeElement.blur();
});
"""


def cell_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    return str(value)


def render(xlsx_path: Path, out_path: Path, title: str | None = None) -> None:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    tabs, panes = [], []

    for i, name in enumerate(wb.sheetnames):
        ws = wb[name]
        rows = [r for r in ws.iter_rows(values_only=True)
                if r and any(c is not None for c in r)]
        on = " class=on" if i == 0 else ""
        pane_on = " on" if i == 0 else ""
        if not rows:
            tabs.append(f'<button onclick="show({i})"{on}>{html.escape(name)}'
                        f'<span class="n">0</span></button>')
            panes.append(f'<section class="pane{pane_on}" data-total="0">'
                         f'<p class="empty">This tab is empty.</p></section>')
            continue

        header, body = rows[0], rows[1:]
        head_html = "".join(f"<th>{html.escape(cell_text(c))}</th>" for c in header)
        body_html = "".join(
            "<tr>" + "".join(f"<td>{html.escape(cell_text(c))}</td>" for c in r)
            + "</tr>" for r in body)

        tabs.append(f'<button onclick="show({i})"{on}>{html.escape(name)}'
                    f'<span class="n">{len(body)}</span></button>')
        panes.append(
            f'<section class="pane{pane_on}" data-total="{len(body)}">'
            f'<div class="tools">'
            f'<input type="search" placeholder="Filter {html.escape(name)}…" '
            f'oninput="filter({i})">'
            f'<span class="count">{len(body)} of {len(body)} rows</span></div>'
            f'<div class="scroll"><table><thead><tr>{head_html}</tr></thead>'
            f'<tbody>{body_html}</tbody></table></div></section>')

    heading = title or xlsx_path.stem.replace("_", " ")
    out_path.write_text(
        "<!doctype html><html><head><meta charset='utf-8'>"
        f"<title>{html.escape(heading)}</title>"
        "<meta name='viewport' content='width=device-width,initial-scale=1'>"
        f"<style>{CSS}</style></head><body>"
        f"<header><h1>{html.escape(heading)}</h1>"
        f"<div class='sub'>{len(wb.sheetnames)} tabs · click a tab, type to "
        f"filter</div></header>"
        f"<nav>{''.join(tabs)}</nav>{''.join(panes)}"
        f"<script>{JS}</script></body></html>",
        encoding="utf-8")
