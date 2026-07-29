"""
Turning a plan into things a human can actually read.

WHY THIS EXISTS
---------------
A plan.json is correct and unreadable. The only thing standing between a
client's twenty years of records and a botched import is one person going
through it and saying "no, that's not her dog". That person needs sheets they
can sort and filter, not a JSON file.

So the plan is rendered three ways, each for a different moment:

  * an XLSX workbook — the archival copy, one tab per kind of decision,
    frozen headers and autofilters so it can be worked through;
  * a self-contained HTML page — because there is no spreadsheet application
    on the machine this runs on, and the data is real names, addresses and
    phone numbers, so it must not go near a hosted viewer (see html_view.py);
  * a folder of CSVs — the machine-readable hand-off, including one file
    shaped for the destination app's own bulk-import wizard.

WHAT DRIVES THE COLUMNS
-----------------------
The workbook is generated FROM the plan, not from a per-client template. The
custom-field columns are discovered by taking the union of every person's
customFields keys, so a client who tracks "Insurance expiry" gets a column
without anyone editing this file.

THE ONE HEADER GOTCHA WORTH KNOWING
-----------------------------------
The destination wizard's auto-mapper sends ANY header containing
"note"/"notes"/"comment" to the DOG's notes field. So the person-level free
text is exported under the header "Background". Rename it and a year of
history about the owner silently lands on the dog.
"""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEAD_FONT = Font(bold=True, color="FFFFFF")
HEAD_FILL = PatternFill("solid", fgColor="1F3864")

# Built-in person columns for the wizard-ready CSV. Custom-field columns are
# appended from the plan itself.
PEOPLE_BASE_HEADERS = ["Client name", "Email", "Phone", "Address",
                       "Dog's name", "Breed"]
BACKGROUND_HEADER = "Background"     # NOT "Notes" — see the module docstring.


# ---------------------------------------------------------------------------
# Workbook
# ---------------------------------------------------------------------------

def _sheet(wb: Workbook, title: str, headers: list[str], rows: list[list],
           widths: list[int]):
    ws = wb.create_sheet(title[:31])
    ws.append(headers)
    for c in ws[1]:
        c.font = HEAD_FONT
        c.fill = HEAD_FILL
        c.alignment = Alignment(vertical="center")
    for r in rows:
        ws.append(["" if v is None else v for v in r])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = (f"A1:{get_column_letter(len(headers))}"
                          f"{max(len(rows) + 1, 2)}")
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)
    return ws


def custom_field_labels(plan: dict) -> list[str]:
    """
    The custom-field columns, in the order the client config DECLARED them,
    with anything else the data turned out to carry appended.

    Declared order first because a column that moves between runs makes two
    review workbooks impossible to diff — and the second review of a client's
    data is usually the one that matters.
    """
    seen: list[str] = [k for k in (plan.get("meta", {}).get("customFields") or [])]
    for p in plan["people"]:
        for k in (p.get("customFields") or {}):
            if k not in seen:
                seen.append(k)
    return seen


def workbook(plan: dict) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    people = plan["people"]
    courses = {c["id"]: c for c in plan["courses"]}
    customs = custom_field_labels(plan)

    # -- People ------------------------------------------------------------
    rows = []
    for p in people:
        dogs = "; ".join(
            " — ".join(x for x in [d["name"] or "(no name)", d["breed"]] if x)
            + (f" (died {d['deceasedAt']})" if d["deceasedAt"] else "")
            for d in p["dogs"])
        rows.append([
            p["id"], p["fullName"], p["firstName"], p["lastName"], p["email"],
            "Yes" if p["emailIsPlaceholder"] else "",
            "; ".join(p["otherEmails"]), p["phone"], p["address"], dogs,
            len(p["enrolments"]), len(p["oneToOneConsults"]),
            *[p["customFields"].get(k, "") for k in customs],
            "; ".join(p["accessCodes"]),
            "; ".join(a.get("raw", "") for a in p["afRefs"]),
            "Yes" if p["isSubscriber"] else "",
            "Yes" if p["waitlist"] else "",
            ", ".join(p["sources"]), "; ".join(p["flags"]), p["notes"],
        ])
    _sheet(wb, "People",
           ["ID", "Full name", "First name", "Last name", "Email",
            "Placeholder email?", "Other emails", "Phone", "Address", "Dogs",
            "Classes", "1:1 consults", *customs, "Access codes", "AF refs",
            "Subscriber", "Waitlist", "Where they came from", "Flags", "Notes"],
           rows,
           [8, 24, 14, 16, 32, 10, 30, 16, 34, 34, 8, 10] + [16] * len(customs)
           + [20, 12, 11, 10, 26, 34, 60])

    # -- Courses -----------------------------------------------------------
    rows = []
    for c in plan["courses"]:
        rows.append([
            c["id"], c["name"], c["level"],
            "From the spreadsheet" if c["origin"] == "spreadsheet" else "NEW — invented",
            c.get("nickname") or "", c.get("day") or "", c.get("date") or "",
            c.get("month") or "", c.get("year") or "", c.get("time") or "",
            c.get("location") or "", "Yes" if c.get("priorProvider") else "",
            c.get("enrolmentCount", 0), "; ".join(c.get("flags") or []),
            c.get("headerRaw") or "",
        ])
    rows.sort(key=lambda r: (r[3], r[2] or "", r[1]))
    _sheet(wb, "Courses",
           ["ID", "Class name", "Level", "Where it came from", "Nickname",
            "Day", "Date", "Month", "Year", "Time", "Location",
            "Previous trainer", "People enrolled", "Notes for you",
            "Original spreadsheet header"],
           rows, [8, 46, 16, 20, 22, 10, 14, 10, 8, 18, 20, 14, 14, 50, 44])

    # -- Enrolments --------------------------------------------------------
    rows = []
    for p in people:
        for e in p["enrolments"]:
            rows.append([
                p["id"], p["fullName"], p["email"], e["courseId"],
                e["courseName"], e["level"] or "",
                "From the spreadsheet" if e["courseOrigin"] == "spreadsheet"
                else "NEW — invented",
                e["dog"], "Yes" if e["didNotAttend"] else "",
                "Yes" if e["priorProvider"] else "",
                ", ".join(e["sources"]), "; ".join(e["rawRefs"]),
                ", ".join(e["sourceRows"]),
            ])
    rows.sort(key=lambda r: (r[4], r[1]))
    _sheet(wb, "Enrolments",
           ["Person ID", "Person", "Email", "Class ID", "Class", "Level",
            "Class origin", "Dog", "Did not attend", "Previous trainer",
            "Where it came from", "Original text", "Spreadsheet rows"],
           rows, [10, 24, 30, 9, 44, 16, 20, 16, 14, 16, 20, 44, 16])

    # -- 1:1 consults ------------------------------------------------------
    rows = []
    for p in people:
        for c in p["oneToOneConsults"]:
            rows.append([p["id"], p["fullName"], p["email"], c["code"],
                         c["service"], c.get("month") or "", c.get("year") or "",
                         c.get("raw") or ""])
    rows.sort(key=lambda r: (r[1],))
    _sheet(wb, "1-1 Consults",
           ["Person ID", "Person", "Email", "Code", "Service", "Month", "Year",
            "Original text"], rows, [10, 24, 30, 8, 32, 12, 8, 60])

    # -- Professionals / role lists ---------------------------------------
    if plan["professionals"]:
        rows = [[r["personId"], r.get("name", ""), r.get("nameInPlan", ""),
                 r.get("business", ""), r.get("email", ""), r.get("phone", ""),
                 r.get("role", ""),
                 {"email": "Matched an existing client by email",
                  "name": "Matched an existing client by name",
                  "new": "New person — not a client"}.get(r.get("matchedBy"),
                                                          r.get("matchedBy", "")),
                 r.get("classes", 0), "; ".join(r.get("flags") or [])]
                for r in plan["professionals"]]
        _sheet(wb, "Professionals",
               ["Person ID", "Name on the list", "Name in the plan", "Business",
                "Email", "Phone", "Role", "How it was matched", "Classes",
                "Flags"], rows, [10, 24, 24, 34, 34, 18, 16, 34, 9, 40])

    # -- Subscribers -------------------------------------------------------
    by_id = {p["id"]: p for p in people}
    rows = []
    for s in plan["subscribers"]:
        person = by_id.get(s["personId"], {})
        rows.append([
            s["personId"], s["name"], s["email"], person.get("fullName", ""),
            {"email": "Already a client (matched on email)",
             "name": "Already a client (matched on name)",
             "new": "Mailing list only — never took a class"}.get(
                s["linkedBy"], s["linkedBy"]),
            len(person.get("enrolments", [])),
            len(person.get("oneToOneConsults", [])),
            s["acceptsMarketing"], s["createdOn"],
        ])
    rows.sort(key=lambda r: (r[4], r[1]))
    _sheet(wb, "Subscribers",
           ["Person ID", "Name on the list", "Email", "Name in the plan",
            "Are they a client?", "Classes", "1:1 consults",
            "Accepts marketing", "Subscribed on"],
           rows, [10, 24, 34, 24, 38, 9, 12, 18, 32])

    # -- Overrides + conflicts --------------------------------------------
    rows = []
    top = plan["meta"].get("rules", {}).get("precedence", [])
    winner = top[0] if top else ""
    for o in plan["fieldOverrides"]:
        rows.append([
            o["personId"], o["person"], o["field"], o["kept"], o["keptFrom"],
            o["discarded"], o["discardedFrom"],
            f"{o['keptFrom']} beats {o['discardedFrom']}"
            if o["keptFrom"] != winner else f"{winner} wins",
        ])
    for j in plan["junkDogNameOverrides"]:
        rows.append(["", j["person"], "dog.name",
                     "(ignored — a real name was taken from another file where "
                     "there was one)", "", j["value"], j["where"],
                     f"Not a dog's name ({j['reason']})"])
    _sheet(wb, "Overrides-and-Conflicts",
           ["Person ID", "Person", "What", "Value kept", "Kept from",
            "Value set aside", "Set aside from", "Why"],
           rows, [10, 24, 16, 40, 16, 40, 26, 40])

    # -- Needs review ------------------------------------------------------
    rows = [[n["kind"], n.get("who", ""), n["detail"]] for n in plan["needsReview"]]
    rows.sort(key=lambda r: (r[0], r[1]))
    _sheet(wb, "Needs-Review", ["What to look at", "Who", "Detail"], rows,
           [40, 26, 110])

    return wb


# ---------------------------------------------------------------------------
# CSVs
# ---------------------------------------------------------------------------

def _full_address(p: dict) -> str:
    parts = [p.get("address") or "", p.get("city") or "", p.get("postcode") or ""]
    return ", ".join(x.strip() for x in parts if x and x.strip())


def people_rows(plan: dict, customs: list[str]):
    """One row per person × dog. A person with no dog still gets a row."""
    csv_cfg = (plan.get("meta") or {}).get("csv") or {}
    company_to = csv_cfg.get("companyTo", "Business")
    once_only = set(csv_cfg.get("oncePerPerson") or ["Credit"])

    for p in plan["people"]:
        cf = dict(p.get("customFields") or {})
        # The destination wizard has no Company field of its own, so a
        # person's company fills the mapped custom column when it is empty.
        if company_to and p.get("company") and not cf.get(company_to):
            cf[company_to] = p["company"]
        # A placeholder email is invented, not evidence. Leaving it blank makes
        # the destination mint its own non-merging placeholder rather than
        # importing a fake address as though it were real.
        email = "" if p.get("emailIsPlaceholder") else (p.get("email") or "")
        base = {
            "Client name": p.get("fullName") or " ".join(
                x for x in [p.get("firstName"), p.get("lastName")] if x).strip(),
            "Email": email,
            "Phone": p.get("phone") or "",
            "Address": _full_address(p),
            BACKGROUND_HEADER: p.get("notes") or "",
            **{k: cf.get(k, "") for k in customs},
        }
        dogs = p.get("dogs") or []
        if not dogs:
            yield {**base, "Dog's name": "", "Breed": ""}
            continue
        for i, dog in enumerate(dogs):
            row = {**base, "Dog's name": dog.get("name") or "",
                   "Breed": dog.get("breed") or ""}
            if i:
                # Same person, extra dog: the wizard merges on email, so
                # repeating the owner columns is harmless. The long free-text
                # ones are not repeated — three copies of a paragraph of notes
                # in a three-dog family is noise a reviewer has to read past.
                row[BACKGROUND_HEADER] = ""
                for k in customs:
                    if k in once_only:
                        row[k] = ""
            yield row


def write_csv(path: Path, headers: list[str], rows: list[dict]) -> int:
    with path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=headers, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)
    return len(rows)


def csvs(plan: dict, out_dir: Path) -> list[tuple[str, int]]:
    out_dir.mkdir(parents=True, exist_ok=True)
    customs = custom_field_labels(plan)
    written: list[tuple[str, int]] = []

    headers = PEOPLE_BASE_HEADERS + customs + [BACKGROUND_HEADER]
    rows = list(people_rows(plan, customs))
    written.append(("01_people_and_dogs.csv",
                    write_csv(out_dir / "01_people_and_dogs.csv", headers, rows)))

    written.append(("02_courses.csv", write_csv(
        out_dir / "02_courses.csv",
        ["id", "origin", "level", "name", "nickname", "date", "month", "year",
         "time", "location", "priorProvider", "enrolmentCount"],
        plan["courses"])))

    enrolments = [
        {"person": p.get("fullName"),
         "email": "" if p.get("emailIsPlaceholder") else p.get("email"),
         "courseId": e.get("courseId"), "course": e.get("courseName"),
         "level": e.get("level"), "courseOrigin": e.get("courseOrigin"),
         "dog": e.get("dog"), "priorProvider": e.get("priorProvider"),
         "sources": ", ".join(e.get("sources") or [])}
        for p in plan["people"] for e in (p.get("enrolments") or [])
    ]
    written.append(("03_enrolments.csv", write_csv(
        out_dir / "03_enrolments.csv",
        ["person", "email", "courseId", "course", "level", "courseOrigin",
         "dog", "priorProvider", "sources"], enrolments)))

    consults = [
        {"person": p.get("fullName"),
         "email": "" if p.get("emailIsPlaceholder") else p.get("email"),
         **({"detail": c} if isinstance(c, str) else c)}
        for p in plan["people"] for c in (p.get("oneToOneConsults") or [])
    ]
    consult_headers = ["person", "email"] + sorted(
        {k for c in consults for k in c if k not in ("person", "email")})
    written.append(("04_one_to_one_consults.csv", write_csv(
        out_dir / "04_one_to_one_consults.csv", consult_headers, consults)))

    written.append(("05_subscribers.csv", write_csv(
        out_dir / "05_subscribers.csv",
        ["email", "name", "linkedBy", "acceptsMarketing", "createdOn"],
        plan["subscribers"])))
    written.append(("06_needs_review.csv", write_csv(
        out_dir / "06_needs_review.csv", ["kind", "who", "detail"],
        plan["needsReview"])))
    written.append(("07_field_overrides.csv", write_csv(
        out_dir / "07_field_overrides.csv",
        ["person", "field", "kept", "keptFrom", "discarded", "discardedFrom"],
        plan["fieldOverrides"])))
    written.append(("08_junk_dog_names.csv", write_csv(
        out_dir / "08_junk_dog_names.csv", ["person", "where", "value", "reason"],
        plan["junkDogNameOverrides"])))
    return written
