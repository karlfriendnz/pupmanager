#!/usr/bin/env python3
"""
PRIME PUPS / "Journey Dog Training" — IMPORT PLAN BUILDER.

Merges the four extracted sources into ONE reviewable import plan.
It writes two files and TOUCHES NO DATABASE:

    /Users/karl/Desktop/Temp/import_plan.json          the machine-readable plan
    /Users/karl/Desktop/Temp/import_plan_review.xlsx   the human review workbook

Sources (in precedence order — see MERGE RULES below):

    prime_extracted.json    the spreadsheet. THE SPINE. 354 clients / 495
                            enrolments / 63 courses / 100 waitlist.
    contacts_parsed.json    519 parsed Outlook contacts (scripts/prime/parse_contacts.py).
    professionals.csv       26 pet-industry professionals.
    profiles.csv            583 Squarespace subscribers.

MERGE RULES (decided — do not change without asking):
  1 Spreadsheet wins every field-level disagreement; contacts fills the gaps.
  2 Junk spreadsheet dog names ("Pulled out", "Class notes please", "Mixed",
    bare breed names) are treated as EMPTY so a real name from contacts wins.
  3 Match on normalised email, else normalised full name. Contacts with no
    email get a placeholder address and can never merge with anyone.
  4 Professionals merge into the client they already are — but two
    professionals sharing one email (Ryan & Annie Black) stay two people.
  5 Volunteers / interns   -> custom field  Volunteer        = "Yes"
  6 Previous-trainer history is imported -> custom field  Previous trainer = "Yes"
  7 Every course referenced anywhere exists in the plan; ones the spreadsheet
    never had are marked  origin: "invented".
  8 Credit notes are kept verbatim in the custom field  Credit.
  9 Door/lockbox codes live in their own accessCodes key, never in notes.
 10 Dogs marked PTS get deceasedAt set.
 11 IN-HOME / DTP / PFP / DTS / ADT / OC / PFS / TW refs become oneToOneConsults,
    not class enrolments.
 12 All 583 subscribers are in the plan, including the ones who never took a class.
 13 Nothing is silently dropped — unclassified note lines survive into notes.

Run:  python3 scripts/prime/build-plan.py
"""

from __future__ import annotations

import csv
import hashlib
import importlib.util
import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

# ── Paths ────────────────────────────────────────────────────────────────────

HERE = Path(__file__).resolve().parent
PRIME_JSON = Path("/Users/karl/Desktop/Temp/prime_extracted.json")
CONTACTS_JSON = Path("/Users/karl/Desktop/Temp/contacts_parsed.json")
PROFESSIONALS_CSV = Path("/Users/karl/Downloads/professionals.csv")
PROFILES_CSV = Path("/Users/karl/Downloads/profiles.csv")

OUT_DIR = Path("/Users/karl/Desktop/Temp")
OUT_JSON = OUT_DIR / "import_plan.json"
OUT_XLSX = OUT_DIR / "import_plan_review.xlsx"

# ── Shared normalisation, borrowed from the parser that produced contacts ─────
# Imported rather than re-implemented so the two scripts can never drift apart.

_spec = importlib.util.spec_from_file_location("prime_parse_contacts",
                                               HERE / "parse_contacts.py")
pc = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(pc)  # module level only defines constants/functions

norm_email = pc.norm_email
norm_name = pc.norm_name
norm_nickname = pc.norm_nickname
norm_breed = pc.norm_breed
breeds_agree = pc.breeds_agree
addresses_agree = pc.addresses_agree
digits = pc.digits
resolve_nickname = pc.resolve_nickname

# ── Constants ────────────────────────────────────────────────────────────────

CODE_LEVEL = {"SP": "School Pups", "ELP": "Early Learning", "TT": "Top Teens"}
LEVEL_CODE = {v: k for k, v in CODE_LEVEL.items()}

SERVICE_NAMES = {
    "DTS": "Dog Training Session (1:1)",
    "DTP": "Dog Training Package (1:1)",
    "PFP": "Puppy Foundation Package (1:1)",
    "PFS": "Puppy Foundation Session (1:1)",
    "TW": "Training Walk",
    "ADT": "Adolescent Dog Training (1:1)",
    "OC": "One-off Consult",
}

# Source priority. Higher wins a straight disagreement.
PRIORITY = {
    "spreadsheet": 40,
    "waitlist": 35,       # also the spreadsheet, but the messier sheet
    "contacts": 30,
    "professionals": 20,
    "subscribers": 10,
}

# Dog-name values that are not dog names.
DOG_NAME_BLOCKLIST = {
    "pulled out", "pulled  out", "class notes please", "classnotes please",
    "mixed", "n a", "na", "tbc", "unknown", "none", "no dog", "x",
}

# Breed vocabulary, split in two so a dog genuinely called "Blue" survives.
#
# STRONG — a word that is a breed and effectively never a dog's given name.
#          One of these on its own means the cell holds a breed, not a name.
# WEAK   — a word that only reads as a breed next to a strong one
#          ("Blue Heeler" is a breed; "Blue" is a perfectly good dog).
STRONG_BREED = {
    "terrier", "retriever", "shepherd", "shepard", "spaniel", "hound", "heeler",
    "collie", "poodle", "labradoodle", "cavoodle", "cavapoo", "pomapoo",
    "spoodle", "schnoodle", "groodle", "goldendoodle", "cockapoo", "mastiff",
    "bulldog", "dogue", "pointer", "setter", "husky", "corgi", "schnauzer",
    "dachshund", "beagle", "labrador", "whippet", "greyhound", "ridgeback",
    "malamute", "akita", "chihuahua", "rottweiler", "doberman", "dobermann",
    "vizsla", "weimaraner", "staffy", "staffie", "staffordshire", "kelpie",
    "huntaway", "samoyed", "shiba", "papillon", "maltese", "bichon",
    "newfoundland", "leonberger", "wolfhound", "deerhound", "saluki", "basenji",
    "dane", "pomeranian", "havanese", "keeshond", "schipperke", "crossbreed",
}
WEAK_BREED = {
    "blue", "red", "black", "white", "tan", "german", "golden", "american",
    "english", "french", "border", "bernese", "mountain", "jack", "russell",
    "fox", "foxy", "bull", "pit", "pitbull", "cocker", "springer", "shih",
    "tzu", "pyrenees", "inu", "miniature", "mini", "standard", "toy", "cross",
    "x", "mix", "mixed", "smooth", "wire", "haired", "shorthaired",
    "longhaired", "working", "show", "heading", "rough", "short", "long",
    "dog", "dogs", "pup", "puppy", "rescue", "lab", "boxer", "pug", "frise",
    "beardie", "bearded", "rottie", "breed",
}

PTS_DATE_RE = re.compile(r"(\d{1,2})[/\-.](\d{1,2})[/\-.](\d{2,4})")

# A "name" from the spreadsheet that is not a person's name.
JUNK_PERSON_RE = re.compile(
    r"^(rehomed|pulled out|cancelled|class notes|n/a|unknown|tbc|-+)$", re.I)

# Rule 5 — the spreadsheet marks its helpers in the notes ("B:Volunteer",
# "F:INTERN"); the contacts file marks them in its own roles list.
ROLE_WORD_RE = re.compile(r"\b(intern|volunteer|on placement)\b", re.I)


def die(msg: str) -> None:
    print(f"ERROR: {msg}", file=sys.stderr)
    raise SystemExit(1)


# ── Small helpers ────────────────────────────────────────────────────────────

def clean(v) -> str:
    return (v or "").strip() if isinstance(v, str) else ("" if v is None else str(v).strip())


def name_tokens(v: str) -> set[str]:
    return {t for t in norm_name(v).split() if t and t != "and"}


def names_compatible(a: str, b: str) -> bool:
    """
    True when one name is the same as, or a subset of, the other — tolerating
    a one-letter spelling difference per word, so "Sara Elliott-Warren" and
    "Sara Elliot-Warren" are one person while "Ryan Black" and "Annie Black"
    stay two.
    """
    ta, tb = name_tokens(a), name_tokens(b)
    if not ta or not tb:
        return False
    if ta == tb or ta <= tb or tb <= ta:
        return True

    def covered(small: set[str], big: set[str]) -> bool:
        return all(
            t in big or any(len(t) > 3 and pc._edit_distance_le1(t, u) for u in big)
            for t in small)

    return covered(ta, tb) or covered(tb, ta)


def placeholder_email(seed: str) -> str:
    """Matches src/app/api/clients/route.ts. Deterministic so re-runs are stable."""
    h = hashlib.sha256(f"prime::{seed}".encode()).hexdigest()[:16]
    return f"noemail-{h}@no-email.pupmanager.app"


def is_placeholder(email: str) -> bool:
    return email.endswith("@no-email.pupmanager.app")


def split_name(full: str) -> tuple[str, str]:
    parts = clean(full).split()
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], " ".join(parts[1:])


def month_num(m: str | None) -> int:
    if not m:
        return 0
    m = m.strip().lower()[:3]
    return {"jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6, "jul": 7,
            "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12}.get(m, 0)


def values_agree(field: str, a: str, b: str) -> bool:
    """Are two values for the same field the same thing said differently?"""
    if not a or not b:
        return True
    if field == "phone":
        da, db = digits(a), digits(b)
        return da == db or da.endswith(db) or db.endswith(da)
    if field == "email":
        return norm_email(a) == norm_email(b)
    if field in ("address",):
        return addresses_agree(a, b)
    if field in ("name", "firstName", "lastName", "city"):
        return norm_name(a) == norm_name(b)
    return a.strip().lower() == b.strip().lower()


# ── Load ─────────────────────────────────────────────────────────────────────

def load_all() -> tuple[dict, list[dict], list[dict], list[dict]]:
    for p in (PRIME_JSON, CONTACTS_JSON, PROFESSIONALS_CSV, PROFILES_CSV):
        if not p.exists():
            die(f"missing input {p}")
    prime = json.loads(PRIME_JSON.read_text())
    contacts_doc = json.loads(CONTACTS_JSON.read_text())
    contacts = [c for c in contacts_doc["contacts"] if not c.get("isBlankRow")]
    with PROFESSIONALS_CSV.open(newline="", encoding="utf-8-sig") as f:
        professionals = list(csv.DictReader(f))
    with PROFILES_CSV.open(newline="", encoding="utf-8-sig") as f:
        profiles = list(csv.DictReader(f))
    return prime, contacts, professionals, profiles


# ── Dog-name junk detection (rule 2) ─────────────────────────────────────────

class DogNameJudge:
    """Decides whether a spreadsheet dog-name cell actually holds a dog's name."""

    def __init__(self, prime: dict, contacts: list[dict]) -> None:
        vocab: set[str] = set()
        for e in prime.get("enrolments", []):
            self._add_breed(vocab, e.get("breed"))
        for c in contacts:
            for d in c.get("dogs", []):
                self._add_breed(vocab, d.get("breed"))
        self.vocab = vocab
        self.overrides: list[dict] = []
        self._seen: set[tuple] = set()

    @staticmethod
    def _add_breed(vocab: set[str], breed: str | None) -> None:
        b = norm_breed(breed)
        if not b:
            return
        vocab.add(b)
        for w in b.split():
            if len(w) > 4:
                vocab.add(w)

    def junk_reason(self, raw: str | None) -> str | None:
        s = clean(raw)
        if not s:
            return "empty"
        base = re.sub(r"\(.*?\)", " ", s).strip()          # drop "(F)", "(M)"
        key = norm_name(base)
        if not key:
            return "empty"
        if key in DOG_NAME_BLOCKLIST:
            return "blocklist"
        toks = key.split()
        all_breed_words = all(t in STRONG_BREED or t in WEAK_BREED for t in toks)
        has_strong = any(t in STRONG_BREED for t in toks)
        # One word only counts if it is unmistakably a breed — a dog really
        # can be called "Blue", but nobody calls one "Doberman".
        if len(toks) == 1:
            return "breed-name-only" if has_strong else None
        # Two or more words that are all breed vocabulary ("Blue Heeler",
        # "Lab mix"). "Frankie Bear" and "Tim Tam" are untouched.
        if all_breed_words:
            return "breed-name-only"
        # Or the whole string is a breed we have seen in the breed columns.
        if has_strong and norm_breed(base) in self.vocab:
            return "breed-name-only"
        return None

    def usable(self, raw: str | None, *, owner: str = "", where: str = "",
               person: "Person | None" = None) -> str:
        """Return the name if it is one, else "" — logging the override."""
        reason = self.junk_reason(raw)
        if reason is None:
            return clean(raw)
        value = clean(raw)
        if reason != "empty":
            key = (owner, value.lower(), reason)
            if key not in self._seen:
                self._seen.add(key)
                self.overrides.append({
                    "person": owner,
                    "where": where,
                    "value": value,
                    "reason": reason,
                })
                # Rule 13 — the text still means something, so it survives.
                if person is not None:
                    person.add_note(f'Dog column on {where} said: "{value}"',
                                    "spreadsheet")
        return ""


# ── Course catalogue ─────────────────────────────────────────────────────────

class CourseCatalogue:
    """Every course anyone refers to — from the spreadsheet, or invented."""

    def __init__(self, prime: dict) -> None:
        self.courses: dict[str, dict] = {}
        self.by_nick: dict[str, list[dict]] = defaultdict(list)
        self.needs_review: list[dict] = []

        for c in prime.get("courses", []):
            key = norm_nickname(c.get("nickname"))
            cid = f"S{c['id']}"
            rec = {
                "id": cid,
                "origin": "spreadsheet",
                "spreadsheetCourseId": c["id"],
                "level": c.get("sheet"),
                "code": LEVEL_CODE.get(c.get("sheet") or "", ""),
                "name": self._spreadsheet_name(c),
                "nickname": c.get("nickname"),
                "nicknameKey": key,
                "day": c.get("day"),
                "date": c.get("date"),
                "month": None,
                "year": None,
                "time": c.get("time"),
                "location": c.get("location"),
                "priorProvider": False,
                "headerRaw": c.get("headerRaw"),
                "sourceRefs": [],
                "flags": [],
            }
            self.courses[cid] = rec
            if key:
                self.by_nick[key].append(rec)

        self.nick_keys = list(self.by_nick)
        self._invented: dict[tuple, dict] = {}
        self._invented_seq = 0

    @staticmethod
    def _spreadsheet_name(c: dict) -> str:
        nick = clean(c.get("nickname"))
        when = " ".join(x for x in [clean(c.get("day")), clean(c.get("date"))] if x)
        if nick:
            return f"{c.get('sheet')} — {nick}"
        return f"{c.get('sheet')} — {when}" if when else f"{c.get('sheet')} — Class {c['id']}"

    # -- matching ------------------------------------------------------------

    def match_ref(self, ref: dict, who: str) -> dict | None:
        """Resolve a contacts course reference onto a spreadsheet course."""
        key = ref.get("nicknameKey")
        if not key:
            return None
        matched, how = resolve_nickname(key, self.nick_keys)
        if not matched:
            return None
        cands = self.by_nick[matched]
        want = CODE_LEVEL.get(ref.get("code") or "")
        if len(cands) > 1:
            same = [c for c in cands if c["level"] == want]
            if len(same) == 1:
                return same[0]
            self.needs_review.append({
                "kind": "Ambiguous course nickname",
                "who": who,
                "detail": (f"'{ref.get('raw')}' matches {len(cands)} spreadsheet "
                           f"classes ({', '.join(c['name'] for c in cands)}). "
                           f"Used the first."),
            })
            return cands[0]
        course = cands[0]
        if want and course["level"] != want:
            self.needs_review.append({
                "kind": "Class level disagrees",
                "who": who,
                "detail": (f"Contacts says {ref.get('code')} ({want}) but the "
                           f"spreadsheet class '{course['name']}' is "
                           f"{course['level']}. Kept the spreadsheet class."),
            })
            course["flags"].append(f"level mismatch vs contacts ({ref.get('code')})")
        if how != "exact":
            course["flags"].append(f"matched by {how}: '{ref.get('nickname')}'")
        return course

    # -- invention -----------------------------------------------------------

    @staticmethod
    def _usable_nickname(ref: dict) -> str:
        """A nickname we can name a class after — junk parses excluded."""
        key = clean(ref.get("nicknameKey"))
        raw = clean(ref.get("nickname"))
        if not key or not raw:
            return ""
        junk = {"the", "allsort", "allsorts", "and sps allsort", "sp", "sps",
                "elp", "tt", "journey"}
        if key in junk:
            return ""
        toks = [t for t in key.split() if t not in junk]
        return raw if toks else ""

    def invent_for(self, ref: dict) -> dict:
        nick = self._usable_nickname(ref)
        code = clean(ref.get("code")) or "SP"
        level = CODE_LEVEL.get(code, code)
        month, year = clean(ref.get("month")), clean(ref.get("year"))
        prior = bool(ref.get("priorProvider"))
        key = (code, month, year, norm_nickname(nick), prior)
        if key in self._invented:
            return self._invented[key]

        self._invented_seq += 1
        when = " ".join(x for x in [month, year] if x) or "date unknown"
        name = f"{level} — {when}"
        if nick:
            name = f"{level} — {nick} ({when})" if when != "date unknown" else f"{level} — {nick}"
        if prior:
            name += " (previous trainer)"

        cid = f"N{self._invented_seq:03d}"
        rec = {
            "id": cid,
            "origin": "invented",
            "spreadsheetCourseId": None,
            "level": level,
            "code": code,
            "name": name,
            "nickname": nick or None,
            "nicknameKey": norm_nickname(nick) if nick else "",
            "day": None,
            "date": None,
            "month": month or None,
            "year": year or None,
            "time": None,
            "location": None,
            "priorProvider": prior,
            "headerRaw": None,
            "sourceRefs": [],
            "flags": ["no nickname in the contacts note — named from its date"]
                     if not nick else [],
        }
        # Is it plausibly one of the spreadsheet classes whose header was cut off?
        if nick:
            nk = norm_nickname(nick)
            for c in self.courses.values():
                if c["origin"] != "spreadsheet" or c["nicknameKey"]:
                    continue
                header = norm_nickname(c.get("headerRaw") or "")
                if nk and nk in header:
                    rec["probableSpreadsheetMatch"] = c["id"]
                    rec["flags"].append(
                        f"probably the same class as spreadsheet '{c['name']}' "
                        f"(its header row was cut off) — check before importing")
                    self.needs_review.append({
                        "kind": "Invented class may already exist",
                        "who": "",
                        "detail": (f"New class '{name}' looks like spreadsheet class "
                                   f"{c['id']} ('{c['headerRaw']}'). Kept separate "
                                   f"as instructed — merge if she agrees."),
                    })
                    break
        self._invented[key] = rec
        self.courses[cid] = rec
        return rec


# ── Person registry ──────────────────────────────────────────────────────────

class Person:
    __slots__ = ("id", "fields", "provenance", "overrides", "dogs", "enrolments",
                 "consults", "custom", "notes", "access_codes", "af_refs",
                 "waitlist", "sources", "flags", "subscriber", "claimed_by_prof",
                 "prof", "other_emails")

    def __init__(self, pid: str) -> None:
        self.id = pid
        self.fields: dict[str, str] = {}
        self.provenance: dict[str, str] = {}
        self.overrides: list[dict] = []
        self.dogs: list[dict] = []
        self.enrolments: list[dict] = []
        self.consults: list[dict] = []
        self.custom: dict[str, str] = {}
        self.notes: list[str] = []
        self.access_codes: list[str] = []
        self.af_refs: list[dict] = []
        self.waitlist: list[dict] = []
        self.sources: list[str] = []
        self.flags: list[str] = []
        self.subscriber: dict | None = None
        self.claimed_by_prof = False
        self.prof: dict | None = None
        self.other_emails: list[str] = []

    def add_other_email(self, email: str) -> None:
        e = norm_email(email)
        if (e and not is_placeholder(e)
                and e != norm_email(self.fields.get("email", ""))
                and e not in self.other_emails):
            self.other_emails.append(e)

    # -- field precedence ----------------------------------------------------

    def set_field(self, field: str, value, source: str, registry: "Registry") -> None:
        v = clean(value)
        if not v:
            return
        cur = self.fields.get(field, "")
        if not cur:
            self.fields[field] = v
            self.provenance[field] = source
            return
        if values_agree(field, cur, v):
            return
        if field == "email":
            # A placeholder is invented, not evidence. A real address always
            # wins, whichever file it came from, and no conflict is recorded.
            if is_placeholder(cur) and not is_placeholder(v):
                self.fields[field] = v
                self.provenance[field] = source
                return
            if is_placeholder(v):
                return
        cur_src = self.provenance.get(field, "unknown")
        if PRIORITY.get(source, 0) > PRIORITY.get(cur_src, 0):
            kept, kept_src, dropped, dropped_src = v, source, cur, cur_src
            self.fields[field] = v
            self.provenance[field] = source
        else:
            kept, kept_src, dropped, dropped_src = cur, cur_src, v, source
        if field == "email":
            # An address that lost the precedence fight is still a real address.
            self.add_other_email(dropped)
        entry = {
            "personId": self.id,
            "person": self.fields.get("name", ""),
            "field": field,
            "kept": kept,
            "keptFrom": kept_src,
            "discarded": dropped,
            "discardedFrom": dropped_src,
        }
        self.overrides.append(entry)
        registry.overrides.append(entry)

    def add_note(self, line: str, source: str) -> None:
        t = clean(line)
        if not t:
            return
        tagged = f"[{source}] {t}"
        if tagged not in self.notes:
            self.notes.append(tagged)

    def add_source(self, s: str) -> None:
        if s not in self.sources:
            self.sources.append(s)

    def add_flag(self, f: str) -> None:
        if f not in self.flags:
            self.flags.append(f)


class Registry:
    def __init__(self) -> None:
        self.people: list[Person] = []
        self.by_email: dict[str, Person] = {}
        self.by_name: dict[str, list[Person]] = defaultdict(list)
        self.overrides: list[dict] = []
        self.needs_review: list[dict] = []
        self._seq = 0

    def new_person(self, email: str, name: str) -> Person:
        self._seq += 1
        p = Person(f"P{self._seq:04d}")
        self.people.append(p)
        self.index(p, email, name)
        return p

    def index(self, p: Person, email: str, name: str) -> None:
        e = norm_email(email)
        if e and not is_placeholder(e) and e not in self.by_email:
            self.by_email[e] = p
        n = norm_name(name)
        if n and p not in self.by_name[n]:
            self.by_name[n].append(p)

    def find(self, email: str, name: str, *, allow_name: bool = True) -> tuple[Person | None, str]:
        e = norm_email(email)
        if e and not is_placeholder(e) and e in self.by_email:
            return self.by_email[e], "email"
        if allow_name:
            n = norm_name(name)
            if n and self.by_name.get(n):
                return self.by_name[n][0], "name"
        return None, ""

    def review(self, kind: str, who: str, detail: str) -> None:
        self.needs_review.append({"kind": kind, "who": who, "detail": detail})


# ── Dog merging ──────────────────────────────────────────────────────────────

def merge_dog(person: Person, *, name: str, breed: str, source: str,
              deceased_at: str | None = None, note: str = "",
              registry: Registry | None = None) -> None:
    name, breed = clean(name), clean(breed)
    if not name and not breed and not note:
        return

    target = None
    if name:
        for d in person.dogs:
            if d["name"] and norm_name(d["name"]) == norm_name(name):
                target = d
                break
        if target is None:
            # A nameless / junk-named dog whose breed agrees adopts this name.
            for d in person.dogs:
                if not d["name"] and breeds_agree(d.get("breed") or "", breed):
                    target = d
                    break
    else:
        for d in person.dogs:
            if breed and d.get("breed") and breeds_agree(d["breed"], breed):
                target = d
                break
        if target is None and len(person.dogs) == 1 and not person.dogs[0]["name"]:
            target = person.dogs[0]

    if target is None:
        target = {"name": "", "breed": "", "deceasedAt": None, "notes": [],
                  "sources": [], "provenance": {}, "flags": []}
        person.dogs.append(target)

    for field, value in (("name", name), ("breed", breed)):
        if not value:
            continue
        cur = target[field]
        if not cur:
            target[field] = value
            target["provenance"][field] = source
        elif field == "breed" and breeds_agree(cur, value):
            pass
        elif field == "name" and norm_name(cur) == norm_name(value):
            pass
        else:
            cur_src = target["provenance"].get(field, "unknown")
            if PRIORITY.get(source, 0) > PRIORITY.get(cur_src, 0):
                kept, kept_src, dropped, dropped_src = value, source, cur, cur_src
                target[field] = value
                target["provenance"][field] = source
            else:
                kept, kept_src, dropped, dropped_src = cur, cur_src, value, source
            entry = {
                "personId": person.id,
                "person": person.fields.get("name", ""),
                "field": f"dog.{field}",
                "kept": kept,
                "keptFrom": kept_src,
                "discarded": dropped,
                "discardedFrom": dropped_src,
            }
            target["flags"].append(f"{field} disagreed: kept '{kept}' over '{dropped}'")
            person.overrides.append(entry)
            if registry:
                registry.overrides.append(entry)

    if deceased_at and not target["deceasedAt"]:
        target["deceasedAt"] = deceased_at
    if note and note not in target["notes"]:
        target["notes"].append(note)
    if source not in target["sources"]:
        target["sources"].append(source)


def pts_date(note: str | None) -> str | None:
    """'PTS 21/07/25' -> '2025-07-21'. Returns None when no date is readable."""
    if not note:
        return None
    m = PTS_DATE_RE.search(note)
    if not m:
        return None
    d, mo, y = (int(x) for x in m.groups())
    if y < 100:
        y += 2000
    if not (1 <= d <= 31 and 1 <= mo <= 12):
        return None
    return f"{y:04d}-{mo:02d}-{d:02d}"


# ── Enrolment / consult helpers ──────────────────────────────────────────────

def add_enrolment(person: Person, course: dict, source: str, *, raw: str = "",
                  dog: str = "", did_not_attend: bool = False,
                  prior_provider: bool = False, source_row: str = "") -> None:
    for e in person.enrolments:
        if e["courseId"] == course["id"]:
            if source not in e["sources"]:
                e["sources"].append(source)
            if raw and raw not in e["rawRefs"]:
                e["rawRefs"].append(raw)
            if dog and not e["dog"]:
                e["dog"] = dog
            e["didNotAttend"] = e["didNotAttend"] or did_not_attend
            e["priorProvider"] = e["priorProvider"] or prior_provider
            if source_row and source_row not in e["sourceRows"]:
                e["sourceRows"].append(source_row)
            return
    person.enrolments.append({
        "courseId": course["id"],
        "courseName": course["name"],
        "courseOrigin": course["origin"],
        "level": course["level"],
        "dog": clean(dog),
        "didNotAttend": did_not_attend,
        "priorProvider": prior_provider,
        "sources": [source],
        "rawRefs": [raw] if raw else [],
        "sourceRows": [source_row] if source_row else [],
    })
    if source_row and source_row not in course["sourceRefs"]:
        course["sourceRefs"].append(source_row)


def add_consult(person: Person, ref: dict, source: str) -> None:
    code = clean(ref.get("code")).upper()
    rec = {
        "code": code,
        "service": SERVICE_NAMES.get(code, "1:1 / in-home service"),
        "month": ref.get("month"),
        "year": ref.get("year"),
        "raw": ref.get("raw"),
        "source": source,
    }
    if rec not in person.consults:
        person.consults.append(rec)


# ── Build ────────────────────────────────────────────────────────────────────

def build() -> dict:
    prime, contacts, professionals, profiles = load_all()

    judge = DogNameJudge(prime, contacts)
    catalogue = CourseCatalogue(prime)
    reg = Registry()

    stats = Counter()

    # ── 1. The spreadsheet — the spine ──────────────────────────────────────

    enrol_by_row = {e["sourceCode"]: e for e in prime["enrolments"]}
    courses_by_sid = {c["id"]: c for c in prime["courses"]}

    for cl in prime["clients"]:
        name = clean(cl.get("name"))
        email = norm_email(cl.get("email"))
        if not email:
            email = placeholder_email(f"client::{name}::{','.join(cl.get('sourceRows') or [])}")
            stats["placeholderEmails"] += 1
        p = reg.new_person(email, name)
        p.add_source("spreadsheet")
        p.set_field("name", name, "spreadsheet", reg)
        p.set_field("email", email, "spreadsheet", reg)
        p.set_field("phone", cl.get("phone"), "spreadsheet", reg)
        p.set_field("address", cl.get("address"), "spreadsheet", reg)
        if is_placeholder(email):
            p.add_flag("no email in any source — placeholder address generated")
        if cl.get("multipleNamesOnKey"):
            p.add_flag("spreadsheet row held more than one name")
            reg.review("More than one name on the row", name,
                       "The spreadsheet cell held several names — check who this is.")
        if JUNK_PERSON_RE.match(name) or len(name) < 2:
            p.add_flag("name may not be a person's name")
            reg.review("Odd client name", name,
                       f"'{name}' does not look like a person's name — check the "
                       f"spreadsheet rows {', '.join(cl.get('sourceRows') or [])}.")
        if name and len(name.split()) == 1:
            p.add_flag("only a first name in the spreadsheet")
        p.add_note(cl.get("notes"), "spreadsheet")
        if ROLE_WORD_RE.search(cl.get("notes") or ""):
            p.custom["Volunteer"] = "Yes"
            stats["volunteersFromSpreadsheet"] += 1

        # Dogs + enrolments come off the enrolment rows (they carry the breed).
        for row in cl.get("sourceRows") or []:
            en = enrol_by_row.get(row)
            if not en:
                continue
            dog_name = judge.usable(en.get("dog"), owner=name, person=p,
                                    where=f"spreadsheet row {row}")
            merge_dog(p, name=dog_name, breed=clean(en.get("breed")),
                      source="spreadsheet", registry=reg,
                      note=f"Age at booking: {clean(en.get('age'))}" if clean(en.get("age")) else "")
            course = courses_by_sid.get(en["courseId"])
            if course:
                add_enrolment(p, catalogue.courses[f"S{course['id']}"], "spreadsheet",
                              dog=dog_name, source_row=row)
        # Any client dog the enrolment rows did not carry.
        for d in cl.get("dogs") or []:
            dn = judge.usable(d, owner=name, person=p,
                              where="the spreadsheet client row")
            if dn:
                merge_dog(p, name=dn, breed="", source="spreadsheet", registry=reg)
        # courseIds with no matching enrolment row (defensive).
        for cid in cl.get("courseIds") or []:
            key = f"S{cid}"
            if key in catalogue.courses and not any(
                    e["courseId"] == key for e in p.enrolments):
                add_enrolment(p, catalogue.courses[key], "spreadsheet")

    stats["spreadsheetClients"] = len(prime["clients"])

    # ── 2. The spreadsheet waitlist ─────────────────────────────────────────

    for w in prime.get("waitlist", []):
        name = clean(w.get("name"))
        email = norm_email(w.get("email"))
        p, how = reg.find(email, name)
        if p is None:
            if not email:
                email = placeholder_email(f"waitlist::{name}::{w.get('sourceRow')}")
                stats["placeholderEmails"] += 1
            p = reg.new_person(email, name)
            stats["waitlistOnlyPeople"] += 1
        else:
            reg.index(p, email, name)
        p.add_source("waitlist")
        p.set_field("name", name, "waitlist", reg)
        p.set_field("email", email, "waitlist", reg)
        p.set_field("phone", w.get("phone"), "waitlist", reg)
        p.set_field("address", w.get("address"), "waitlist", reg)
        p.waitlist.append({
            "category": w.get("category"),
            "sourceRow": w.get("sourceRow"),
            "breedAge": w.get("breedAge"),
            "dogRaw": w.get("dog"),
            "notes": w.get("notes"),
        })
        # The waitlist sheet swaps dog/breed columns often — take only what reads
        # like a name, and keep both raw values in the notes either way.
        dog_raw = clean(w.get("dog"))
        breed_raw = clean(w.get("breedAge"))
        dn = judge.usable(dog_raw, owner=name, person=p,
                          where="the waitlist dog column")
        if dn and re.search(r"\d", dn):     # "born 29th August" is not a name
            p.add_note(f"Waitlist dog column: {dog_raw}", "waitlist")
            dn = ""
        if dn or breed_raw:
            merge_dog(p, name=dn, breed=breed_raw if not re.search(r"\d", breed_raw) else "",
                      source="waitlist", registry=reg)
        if breed_raw and re.search(r"\d", breed_raw):
            p.add_note(f"Waitlist breed/age column: {breed_raw}", "waitlist")
        p.add_note(w.get("notes"), "waitlist")
        if w.get("category"):
            p.add_note(f"On the waitlist for: {w['category']}", "waitlist")

    stats["waitlistRows"] = len(prime.get("waitlist", []))

    # ── 3. Contacts ─────────────────────────────────────────────────────────

    for ct in contacts:
        name = clean(ct.get("fullName"))
        email = norm_email(ct.get("email"))
        if not email:
            # Rule 3: a contact with no email gets a placeholder and never
            # merges with anyone. A same-named client is only pointed out.
            stats["contactsWithNoEmail"] += 1
            twin, _ = reg.find("", name)
            email = placeholder_email(f"contact::{name}::{ct.get('rowNumber')}")
            stats["placeholderEmails"] += 1
            p, how = reg.new_person(email, name), ""
            stats["contactOnlyPeople"] += 1
            if twin is not None:
                p.add_flag(f"possible duplicate of {twin.id}")
                reg.review(
                    "Contact with no email — possible duplicate", name,
                    f"This contact card has no email address, so it was kept as "
                    f"its own person ({p.id}) with a placeholder address. "
                    f"'{twin.fields.get('name')}' ({twin.id}, "
                    f"{twin.fields.get('email')}) has the same name and is "
                    f"probably the same person — merge them if you agree.")
        else:
            p, how = reg.find(email, name)
            if p is None:
                p = reg.new_person(email, name)
                stats["contactOnlyPeople"] += 1
            else:
                stats[f"contactsMatchedBy_{how}"] += 1
                if how == "name" and p.fields.get("email") and not values_agree(
                        "email", p.fields["email"], email):
                    reg.review(
                        "Contact matched on name, not email", name,
                        f"The contact card <{email}> was merged into "
                        f"'{p.fields.get('name')}' ({p.id}) whose email is "
                        f"<{p.fields['email']}> — same name, two addresses. Check.")
                reg.index(p, email, name)
        p.add_source("contacts")
        p.set_field("name", name, "contacts", reg)
        p.set_field("firstName", ct.get("firstName"), "contacts", reg)
        p.set_field("lastName", ct.get("lastName"), "contacts", reg)
        p.set_field("email", email, "contacts", reg)
        p.set_field("phone", ct.get("phone"), "contacts", reg)
        p.set_field("city", ct.get("city"), "contacts", reg)
        p.set_field("postcode", ct.get("postcode"), "contacts", reg)
        p.set_field("company", ct.get("company"), "contacts", reg)
        street = clean(ct.get("street"))
        addr = ", ".join(x for x in [street, clean(ct.get("city"))] if x)
        p.set_field("address", addr, "contacts", reg)
        if clean(ct.get("email2")) and norm_email(ct["email2"]) != norm_email(email):
            p.add_other_email(ct["email2"])
            p.add_note(f"Second email on the contact card: {ct['email2']}", "contacts")

        if is_placeholder(p.fields.get("email", "")):
            p.add_flag("no email in any source — placeholder address generated")

        for d in ct.get("dogs") or []:
            dead = bool(d.get("deceased"))
            merge_dog(p, name=clean(d.get("name")), breed=clean(d.get("breed")),
                      source="contacts", registry=reg,
                      deceased_at=(pts_date(d.get("note")) or "unknown") if dead else None,
                      note=clean(d.get("note")))
            if dead:
                stats["deceasedDogs"] += 1
                when = pts_date(d.get("note"))
                reg.review("Dog has died", name,
                           f"{d.get('name') or 'a dog'} is marked "
                           f"'{d.get('note') or 'PTS'}' — imported as died on "
                           f"{when or 'a date we could not read'}. The owner is "
                           f"imported as normal.")

        # A "PTS"/"RIP"-looking word that did NOT produce a deceased dog is worth
        # a human glance — "Rip" is also a perfectly good dog's name.
        if (re.search(r"\bPTS\b|passed away|\bdied\b|\bRIP\b", ct.get("notesRaw") or "",
                      re.I)
                and not any(d.get("deceased") for d in ct.get("dogs") or [])):
            stats["possibleDeceasedToCheck"] += 1
            reg.review(
                "Might mention a dog that has died", name,
                f"The contact card says: "
                f"{(ct.get('notesRaw') or '').splitlines()[0]!r}. No dog was "
                f"marked as died — check whether it should be.")

        for ref in ct.get("courses") or []:
            course = catalogue.match_ref(ref, name)
            if course is None:
                course = catalogue.invent_for(ref)
                stats["contactRefsInvented"] += 1
            else:
                stats["contactRefsMatched"] += 1
            add_enrolment(p, course, "contacts", raw=clean(ref.get("raw")),
                          dog=clean(ref.get("dog")),
                          did_not_attend=bool(ref.get("didNotAttend")),
                          prior_provider=bool(ref.get("priorProvider")))
            if ref.get("comment"):
                p.add_note(f"{ref.get('raw')} — {ref['comment']}", "contacts")

        for ref in ct.get("oneToOne") or []:
            add_consult(p, ref, "contacts")
            stats["consults"] += 1

        for r in ct.get("roles") or []:
            p.custom["Volunteer"] = "Yes"
            stats["volunteerRefsFromContacts"] += 1
            p.add_note(f"Role: {r.get('raw')}", "contacts")

        for cr in ct.get("credits") or []:
            existing = p.custom.get("Credit", "")
            p.custom["Credit"] = f"{existing} | {cr}".strip(" |") if existing else cr

        for ac in ct.get("accessCodes") or []:
            if ac not in p.access_codes:
                p.access_codes.append(ac)

        for af in ct.get("afRefs") or []:
            if af not in p.af_refs:
                p.af_refs.append(af)

        for h in ct.get("addressHints") or []:
            if not p.fields.get("address"):
                p.set_field("address", h, "contacts", reg)
            else:
                p.add_note(f"Address hint: {h}", "contacts")

        # Rule 13 — every line the parser could not classify survives.
        for line in ct.get("unclassified") or []:
            p.add_note(line, "contacts")
        for line in ct.get("freeText") or []:
            p.add_note(line, "contacts")

        if any(x.get("priorProvider") for x in ct.get("courses") or []):
            p.custom["Previous trainer"] = "Yes"
            stats["previousTrainerPeople"] += 1

    # ── 4. Professionals ────────────────────────────────────────────────────

    prof_seen: set[tuple[str, str]] = set()
    prof_rows: list[dict] = []
    for row in professionals:
        name = clean(row.get("Name"))
        email = norm_email(row.get("Email"))
        dedup_key = (email, norm_name(name))
        if dedup_key in prof_seen:
            stats["professionalDuplicateRows"] += 1
            continue
        prof_seen.add(dedup_key)

        p, how = reg.find(email, name)
        # Rule 4: a person can absorb ONE professional. A second professional on
        # the same email (Ryan & Annie Black) becomes their own person.
        if p is not None and p.claimed_by_prof:
            reg.review(
                "Two professionals share one email", name,
                f"{name} shares {email} with "
                f"{p.fields.get('name', 'another professional')}. Kept as two "
                f"separate people, both using that email address.")
            p = None
            how = ""
        if p is not None and how == "email" and not names_compatible(
                name, p.fields.get("name", "")):
            reg.review(
                "Professional's name does not match the client on that email", name,
                f"{name} <{email}> matched the client "
                f"'{p.fields.get('name')}'. Kept as separate people — check.")
            p = None
            how = ""
        if p is None:
            e = email or placeholder_email(f"professional::{name}")
            if not email:
                stats["placeholderEmails"] += 1
            p = reg.new_person(e, name)
            stats["professionalOnlyPeople"] += 1
        else:
            stats[f"professionalsMatchedBy_{how}"] += 1
            reg.index(p, email, name)
        p.claimed_by_prof = True
        p.prof = row
        p.add_source("professionals")
        p.set_field("name", name, "professionals", reg)
        p.set_field("email", email, "professionals", reg)
        p.set_field("phone", row.get("Phone"), "professionals", reg)
        p.custom["Professional"] = "Yes"
        if clean(row.get("Business")):
            p.custom["Business"] = clean(row.get("Business"))
        if clean(row.get("Group")):
            p.add_note(f"Professional group: {row['Group']}", "professionals")
        prof_rows.append({"row": row, "personId": p.id, "matchedBy": how or "new"})

    stats["professionalRows"] = len(professionals)
    stats["professionalPeople"] = len(prof_rows)

    # ── 5. Subscribers ──────────────────────────────────────────────────────

    subscribers: list[dict] = []
    for row in profiles:
        email = norm_email(row.get("Email"))
        first, last = clean(row.get("First Name")), clean(row.get("Last Name"))
        name = " ".join(x for x in [first, last] if x)
        p, how = reg.find(email, name)
        if p is None:
            p = reg.new_person(email, name)
            stats["subscriberOnlyPeople"] += 1
            how = "new"
        else:
            stats[f"subscribersMatchedBy_{how}"] += 1
            if how == "name" and email and norm_email(p.fields.get("email", "")) != email:
                p.add_note(f"Also subscribes with the email {email}", "subscribers")
                reg.review(
                    "Subscriber matched on name, not email", name,
                    f"The mailing list has {name} <{email}> but this person's "
                    f"email is <{p.fields.get('email')}>. Probably the same "
                    f"person with a typo — merged. Check.")
            reg.index(p, email, name)
        p.add_source("subscribers")
        p.set_field("name", name, "subscribers", reg)
        p.set_field("firstName", first, "subscribers", reg)
        p.set_field("lastName", last, "subscribers", reg)
        p.set_field("email", email, "subscribers", reg)
        p.set_field("phone", row.get("Billing Phone Number") or
                    row.get("Shipping Phone Number"), "subscribers", reg)
        addr = ", ".join(x for x in [clean(row.get("Billing Address 1")),
                                     clean(row.get("Billing City"))] if x)
        p.set_field("address", addr, "subscribers", reg)
        p.subscriber = {
            "email": email,
            "createdOn": clean(row.get("Created On")),
            "subscriberSince": clean(row.get("Subscriber Since")),
            "source": clean(row.get("Subscriber Source")),
            "acceptsMarketing": clean(row.get("Accepts Marketing")),
            "orderCount": clean(row.get("Order Count")),
            "totalSpent": clean(row.get("Total Spent")),
        }
        subscribers.append({
            "email": email,
            "name": name,
            "personId": p.id,
            "linkedBy": how,
            "acceptsMarketing": clean(row.get("Accepts Marketing")),
            "createdOn": clean(row.get("Created On")),
        })

    stats["subscriberRows"] = len(profiles)

    # ── 6. Tidy up + derived flags ──────────────────────────────────────────

    for p in reg.people:
        # A record with neither a name nor a breed is not a dog — whatever was
        # written in its cells goes back onto the person so nothing is lost.
        kept = []
        for d in p.dogs:
            if d["name"] or d["breed"]:
                kept.append(d)
                continue
            for line in d["notes"]:
                p.add_note(line, "spreadsheet")
        p.dogs = kept
        if p.enrolments and not p.dogs:
            stats["peopleInAClassWithNoDog"] += 1
            p.add_flag("takes classes but we have no dog for them")
        if is_placeholder(p.fields.get("email", "")):
            p.add_flag("no email in any source — placeholder address generated")
        if not clean(p.fields.get("name", "")):
            p.add_flag("no name in any source")
            reg.review("Person with no name", p.fields.get("email", ""),
                       f"{p.id} came through with an email address "
                       f"({p.fields.get('email')}) but no name at all.")
        if not p.fields.get("firstName") and p.fields.get("name"):
            f, l = split_name(p.fields["name"])
            p.fields.setdefault("firstName", f)
            if l:
                p.fields.setdefault("lastName", l)
        # Spreadsheet-wins can leave a first-name-only person when contacts had
        # the full name — worth her eye, but the rule stands.
        if (p.provenance.get("name") in ("spreadsheet", "waitlist")
                and len(clean(p.fields.get("name", "")).split()) == 1
                and any(o["field"] == "name" and len(o["discarded"].split()) > 1
                        for o in p.overrides)):
            p.add_flag("spreadsheet had only a first name; a fuller name was discarded")
            reg.review("Spreadsheet name may be incomplete", p.fields.get("name", ""),
                       "The spreadsheet holds only a first name and a fuller name "
                       "from the contacts file was set aside (spreadsheet wins).")
        if p.custom.get("Professional") == "Yes" and not p.custom.get("Business"):
            p.add_flag("professional with no business name")
        if p.access_codes:
            stats["peopleWithAccessCodes"] += 1
        if p.custom.get("Credit"):
            stats["peopleWithCredit"] += 1
        if p.custom.get("Volunteer") == "Yes":
            stats["volunteers"] += 1
        if p.consults:
            stats["peopleWithConsults"] += 1
        if not p.enrolments and not p.consults and not p.waitlist:
            stats["peopleWithNoHistory"] += 1
        if p.subscriber and not p.enrolments and not p.consults:
            stats["subscribersWhoNeverTookAClass"] += 1

    # Course enrolment counts.
    counts = Counter()
    for p in reg.people:
        for e in p.enrolments:
            counts[e["courseId"]] += 1
    for cid, c in catalogue.courses.items():
        c["enrolmentCount"] = counts.get(cid, 0)

    reg.needs_review.extend(catalogue.needs_review)

    # The spreadsheet's own unresolved rows must not vanish.
    for nr in prime.get("needsReview", []):
        reg.review("Spreadsheet row could not be read",
                   "",
                   f"{nr.get('sheet')} row {nr.get('row')} — {nr.get('reason')}: "
                   f"{nr.get('raw')}")

    stats["people"] = len(reg.people)
    stats["courses"] = len(catalogue.courses)
    stats["coursesSpreadsheet"] = sum(1 for c in catalogue.courses.values()
                                      if c["origin"] == "spreadsheet")
    stats["coursesInvented"] = sum(1 for c in catalogue.courses.values()
                                   if c["origin"] == "invented")
    stats["enrolments"] = sum(len(p.enrolments) for p in reg.people)
    stats["dogs"] = sum(len(p.dogs) for p in reg.people)
    stats["fieldOverrides"] = len(reg.overrides)
    stats["junkDogNames"] = len(judge.overrides)

    return {
        "registry": reg,
        "catalogue": catalogue,
        "judge": judge,
        "subscribers": subscribers,
        "professionals": prof_rows,
        "stats": stats,
    }


# ── Serialisation ────────────────────────────────────────────────────────────

def person_to_json(p: Person) -> dict:
    return {
        "id": p.id,
        "firstName": p.fields.get("firstName", ""),
        "lastName": p.fields.get("lastName", ""),
        "fullName": p.fields.get("name", ""),
        "email": p.fields.get("email", ""),
        "emailIsPlaceholder": is_placeholder(p.fields.get("email", "")),
        "otherEmails": p.other_emails,
        "phone": p.fields.get("phone", ""),
        "address": p.fields.get("address", ""),
        "city": p.fields.get("city", ""),
        "postcode": p.fields.get("postcode", ""),
        "company": p.fields.get("company", ""),
        "dogs": [{
            "name": d["name"],
            "breed": d["breed"],
            "deceasedAt": d["deceasedAt"],
            "notes": d["notes"],
            "sources": d["sources"],
            "flags": d["flags"],
        } for d in p.dogs],
        "enrolments": p.enrolments,
        "oneToOneConsults": p.consults,
        "customFields": p.custom,
        "notes": "\n".join(p.notes),
        "accessCodes": p.access_codes,
        "afRefs": p.af_refs,
        "waitlist": p.waitlist,
        "isSubscriber": p.subscriber is not None,
        "subscriber": p.subscriber,
        "sources": p.sources,
        "provenance": p.provenance,
        "overrides": p.overrides,
        "flags": p.flags,
    }


def write_json(result: dict) -> None:
    reg: Registry = result["registry"]
    cat: CourseCatalogue = result["catalogue"]
    judge: DogNameJudge = result["judge"]
    plan = {
        "meta": {
            "generatedBy": "scripts/prime/build-plan.py",
            "sources": {
                "spreadsheet": str(PRIME_JSON),
                "contacts": str(CONTACTS_JSON),
                "professionals": str(PROFESSIONALS_CSV),
                "subscribers": str(PROFILES_CSV),
            },
            "rules": {
                "precedence": ["spreadsheet", "waitlist", "contacts",
                               "professionals", "subscribers"],
                "placeholderEmailPattern": "noemail-<16hex>@no-email.pupmanager.app",
            },
            "counts": dict(sorted(result["stats"].items())),
        },
        "people": [person_to_json(p) for p in reg.people],
        "courses": list(cat.courses.values()),
        "subscribers": result["subscribers"],
        "professionals": result["professionals"],
        "junkDogNameOverrides": judge.overrides,
        "fieldOverrides": reg.overrides,
        "needsReview": reg.needs_review,
    }
    OUT_JSON.write_text(json.dumps(plan, indent=2, ensure_ascii=False))
    return plan


# ── Workbook ─────────────────────────────────────────────────────────────────

def write_xlsx(plan: dict, result: dict) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)

    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="1F3864")

    def sheet(title: str, headers: list[str], rows: list[list], widths: list[int]):
        ws = wb.create_sheet(title[:31])
        ws.append(headers)
        for c in ws[1]:
            c.font = head_font
            c.fill = head_fill
            c.alignment = Alignment(vertical="center")
        for r in rows:
            ws.append(["" if v is None else v for v in r])
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(len(rows) + 1, 2)}"
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.alignment = Alignment(vertical="top", wrap_text=True)
        return ws

    people = plan["people"]
    courses = {c["id"]: c for c in plan["courses"]}

    # People -----------------------------------------------------------------
    rows = []
    for p in people:
        dogs = "; ".join(
            " — ".join(x for x in [d["name"] or "(no name)", d["breed"]] if x)
            + (f" (died {d['deceasedAt']})" if d["deceasedAt"] else "")
            for d in p["dogs"])
        rows.append([
            p["id"], p["fullName"], p["firstName"], p["lastName"],
            p["email"], "Yes" if p["emailIsPlaceholder"] else "",
            "; ".join(p["otherEmails"]),
            p["phone"], p["address"],
            dogs,
            len(p["enrolments"]), len(p["oneToOneConsults"]),
            p["customFields"].get("Professional", ""),
            p["customFields"].get("Business", ""),
            p["customFields"].get("Volunteer", ""),
            p["customFields"].get("Previous trainer", ""),
            p["customFields"].get("Credit", ""),
            "; ".join(p["accessCodes"]),
            "; ".join(a.get("raw", "") for a in p["afRefs"]),
            "Yes" if p["isSubscriber"] else "",
            "Yes" if p["waitlist"] else "",
            ", ".join(p["sources"]),
            "; ".join(p["flags"]),
            p["notes"],
        ])
    sheet("People",
          ["ID", "Full name", "First name", "Last name", "Email",
           "Placeholder email?", "Other emails", "Phone", "Address", "Dogs",
           "Classes", "1:1 consults", "Professional", "Business", "Volunteer",
           "Previous trainer", "Credit", "Access codes", "AF refs",
           "Subscriber", "Waitlist", "Where they came from", "Flags", "Notes"],
          rows,
          [8, 24, 14, 16, 32, 10, 30, 16, 34, 34, 8, 10, 12, 24, 10, 14, 34, 20,
           12, 11, 10, 26, 34, 60])

    # Courses ----------------------------------------------------------------
    rows = []
    for c in plan["courses"]:
        rows.append([
            c["id"], c["name"], c["level"],
            "From the spreadsheet" if c["origin"] == "spreadsheet" else "NEW — invented",
            c.get("nickname") or "", c.get("day") or "", c.get("date") or "",
            c.get("month") or "", c.get("year") or "", c.get("time") or "",
            c.get("location") or "",
            "Yes" if c.get("priorProvider") else "",
            c.get("enrolmentCount", 0),
            "; ".join(c.get("flags") or []),
            c.get("headerRaw") or "",
        ])
    rows.sort(key=lambda r: (r[3], r[2] or "", r[1]))
    sheet("Courses",
          ["ID", "Class name", "Level", "Where it came from", "Nickname", "Day",
           "Date", "Month", "Year", "Time", "Location", "Previous trainer",
           "People enrolled", "Notes for you", "Original spreadsheet header"],
          rows,
          [8, 46, 16, 20, 22, 10, 14, 10, 8, 18, 20, 14, 14, 50, 44])

    # Enrolments -------------------------------------------------------------
    rows = []
    for p in people:
        for e in p["enrolments"]:
            c = courses.get(e["courseId"], {})
            rows.append([
                p["id"], p["fullName"], p["email"], e["courseId"], e["courseName"],
                e["level"] or "",
                "From the spreadsheet" if e["courseOrigin"] == "spreadsheet" else "NEW — invented",
                e["dog"], "Yes" if e["didNotAttend"] else "",
                "Yes" if e["priorProvider"] else "",
                ", ".join(e["sources"]), "; ".join(e["rawRefs"]),
                ", ".join(e["sourceRows"]),
            ])
    rows.sort(key=lambda r: (r[4], r[1]))
    sheet("Enrolments",
          ["Person ID", "Person", "Email", "Class ID", "Class", "Level",
           "Class origin", "Dog", "Did not attend", "Previous trainer",
           "Where it came from", "Original text", "Spreadsheet rows"],
          rows,
          [10, 24, 30, 9, 44, 16, 20, 16, 14, 16, 20, 44, 16])

    # 1-1 Consults -----------------------------------------------------------
    rows = []
    for p in people:
        for c in p["oneToOneConsults"]:
            rows.append([p["id"], p["fullName"], p["email"], c["code"],
                         c["service"], c.get("month") or "", c.get("year") or "",
                         c.get("raw") or ""])
    rows.sort(key=lambda r: (r[1],))
    sheet("1-1 Consults",
          ["Person ID", "Person", "Email", "Code", "Service", "Month", "Year",
           "Original text"],
          rows, [10, 24, 30, 8, 32, 12, 8, 60])

    # Professionals ----------------------------------------------------------
    by_id = {p["id"]: p for p in people}
    rows = []
    for pr in plan["professionals"]:
        row, pid = pr["row"], pr["personId"]
        person = by_id.get(pid, {})
        rows.append([
            pid, row.get("Name", ""), person.get("fullName", ""),
            row.get("Business", ""), row.get("Email", ""), row.get("Phone", ""),
            row.get("Group", ""),
            {"email": "Matched an existing client by email",
             "name": "Matched an existing client by name",
             "new": "New person — not a client"}.get(pr["matchedBy"], pr["matchedBy"]),
            len(person.get("enrolments", [])),
            "; ".join(person.get("flags", [])),
        ])
    sheet("Professionals",
          ["Person ID", "Name on the list", "Name in the plan", "Business",
           "Email", "Phone", "Group", "How it was matched", "Classes", "Flags"],
          rows, [10, 24, 24, 34, 34, 18, 22, 34, 9, 40])

    # Subscribers ------------------------------------------------------------
    rows = []
    for s in plan["subscribers"]:
        person = by_id.get(s["personId"], {})
        rows.append([
            s["personId"], s["name"], s["email"], person.get("fullName", ""),
            {"email": "Already a client (matched on email)",
             "name": "Already a client (matched on name)",
             "new": "Mailing list only — never took a class"}.get(s["linkedBy"], s["linkedBy"]),
            len(person.get("enrolments", [])),
            len(person.get("oneToOneConsults", [])),
            s["acceptsMarketing"], s["createdOn"],
        ])
    rows.sort(key=lambda r: (r[4], r[1]))
    sheet("Subscribers",
          ["Person ID", "Name on the list", "Email", "Name in the plan",
           "Are they a client?", "Classes", "1:1 consults", "Accepts marketing",
           "Subscribed on"],
          rows, [10, 24, 34, 24, 38, 9, 12, 18, 32])

    # Overrides and conflicts ------------------------------------------------
    rows = []
    for o in plan["fieldOverrides"]:
        rows.append([
            o["personId"], o["person"], o["field"], o["kept"], o["keptFrom"],
            o["discarded"], o["discardedFrom"],
            "Spreadsheet wins" if o["keptFrom"] in ("spreadsheet", "waitlist")
            else "Best available source wins",
        ])
    for j in plan["junkDogNameOverrides"]:
        rows.append(["", j["person"], "dog.name", "(ignored — a real name was "
                     "taken from the contacts file where there was one)",
                     "contacts", j["value"], j["where"],
                     f"Not a dog's name ({j['reason']})"])
    sheet("Overrides-and-Conflicts",
          ["Person ID", "Person", "What", "Value kept", "Kept from",
           "Value set aside", "Set aside from", "Why"],
          rows, [10, 24, 16, 40, 16, 40, 26, 40])

    # Needs review -----------------------------------------------------------
    rows = [[n["kind"], n.get("who", ""), n["detail"]] for n in plan["needsReview"]]
    rows.sort(key=lambda r: (r[0], r[1]))
    sheet("Needs-Review", ["What to look at", "Who", "Detail"], rows, [40, 26, 110])

    wb.save(OUT_XLSX)


# ── Main ─────────────────────────────────────────────────────────────────────

def main() -> int:
    result = build()
    plan = write_json(result)
    write_xlsx(plan, result)

    s = result["stats"]
    print(f"wrote {OUT_JSON}")
    print(f"wrote {OUT_XLSX}")
    print()
    for k in sorted(s):
        print(f"  {k:38} {s[k]}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
