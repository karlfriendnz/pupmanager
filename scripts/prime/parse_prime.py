#!/usr/bin/env python3
"""
PRIME PUPS / "Journey Dog Training" onboarding data extractor.

LOCAL, READ-ONLY data task. Reads the messy 4-sheet intake spreadsheet and
produces:
  1. A structured JSON extract (courses / enrolments / deduped clients /
     waitlist / needsReview).
  2. A human review workbook (xlsx) for Karl to eyeball before any import.

There is NO database and NO network access here. This script never imports
anything -- it only parses the spreadsheet.

Design notes
------------
* The three class sheets (Early Learning / School Pups / Top Teens) are a
  running historical log of course intakes. The SAME people recur across
  sheets and sections.
* Course boundaries are HEADER ROWS in column A:
    - "STARTS <Weekday> <Day>, <Time> '<Nickname>'"  (STARTS may be lowercase)
    - or a bare "<Weekday> <Day> <Month>, <Time> - <Nickname>" (School Pups
      later rows dropped the "STARTS" prefix).
* Each sheet has a DIFFERENT column order, and Early Learning even shifts
  columns mid-sheet (~row 142 the trainer inserted Puppy Name/Breed/VAXX
  columns). So cells are classified by CONTENT, never by fixed position:
    email  -> contains '@'
    date   -> a real datetime OR a dd/mm[/yy] string  (TESTED BEFORE phone,
              because ISO/slash dates look phone-shaped)
    phone  -> NZ mobile/landline digit pattern
    age    -> contains week / month / yr / year
    breed  -> matches a breed keyword or an "X"/cross marker
    address-> street-type word or a known Bay-of-Plenty suburb
    dog    -> leftover short alpha token (puppy name)
    A      -> ALWAYS the client name
  Anything left over is kept verbatim in `notes` as "colX:<value>" so nothing
  is silently dropped.
* Column G on Early Learning STARTS rows is a LOCATION (purple font) for the
  first ~11 courses; on cancelled/other courses it is a note. We treat a G
  value on an EL header as a location only when it looks like a venue.

Outputs:
  JSON     -> /Users/karl/Desktop/Temp/prime_extracted.json
  Workbook -> /Users/karl/Desktop/Temp/PRIME_PUPS_import_review_v3.xlsx
"""

import json
import re
import datetime
from collections import OrderedDict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SOURCE = "/Users/karl/Downloads/_PRIME PUPS.xlsx"
JSON_OUT = "/Users/karl/Desktop/Temp/prime_extracted.json"
XLSX_OUT = "/Users/karl/Desktop/Temp/PRIME_PUPS_import_review_v3.xlsx"

SHEET_PREFIX = {
    "Early Learning": "Ea",
    "School Pups": "Sc",
    "Top Teens": "Tt",
}
CLASS_SHEETS = ["Early Learning", "School Pups", "Top Teens"]

MAX_COL = 13  # A..L (col 12) is enough; scan one extra for safety

# ---------------------------------------------------------------------------
# Content classifiers
# ---------------------------------------------------------------------------

WEEKDAY = r"(monday|tuesday|wednesday|thursday|friday|saturday|sunday)"
MONTH = (r"(january|february|march|april|may|june|july|august|september|"
         r"october|november|december)")
TIME_RE = re.compile(r"\d{1,2}\s*[:.]?\d{0,2}\s*(am|pm)", re.I)

# A course header: optional "STARTS", then a weekday, and it must carry a time.
HEADER_RE = re.compile(
    r"^\s*(starts\s+)?" + WEEKDAY + r"\b.*?\d", re.I)

DATE_STR_RE = re.compile(r"^\s*\d{1,2}[/.]\d{1,2}([/.]\d{2,4})?\s*$")
# a date embedded (e.g. "13/12/24") - used only as a hint, not primary

PHONE_RE = re.compile(r"^[\s+]*\(?0?\d[\d\s/()+-]{6,}$")

AGE_RE = re.compile(r"\b(week|weeks|wk|wks|month|months|mth|mths|"
                    r"yr|yrs|year|years)\b", re.I)

SUBURBS = [
    "papamoa", "tauranga", "matua", "otumoetai", "pyes pa", "bethlehem",
    "mount maunganui", "mount", "mt maunganui", "katikati", "welcome bay",
    "ohauiti", "brookfield", "bellevue", "greerton", "judea", "whakamarama",
    "te puna", "omokoroa", "maungatapu", "paengaroa", "tauriko", "hairini",
    "poike", "gate pa", "cherrywood", "athenree", "tahawai", "matamata",
    "putararu", "lower kaimai", "rd1", "rd 1", "rd3", "rd4", "rd7", "rd8",
]
STREET_RE = re.compile(
    r"\b(road|rd|street|st|ave|avenue|drive|dr|place|pl|crescent|cres|lane|"
    r"terrace|tce|court|ct|way|grove|blvd|boulevard|close|dell|vista|"
    r"heights|key|mews|parade|highway|rise|hau|cutting)\b", re.I)

BREED_WORDS = [
    "labrador", "retriever", "poodle", "spaniel", "collie", "terrier",
    "doodle", "cavoodle", "schnauzer", "schnoodle", "beagle", "vizsla",
    "mastiff", "chihuahua", "rottweiler", "dachshund", "bulldog", "shepherd",
    "gsd", "gsp", "wss", "groodle", "labradoodle", "goldendoodle", "husky",
    "dalmation", "dalmatian", "boxer", "whippet", "staffy", "staffordshire",
    "heading dog", "huntaway", "foxy", "fox terrier", "papillon", "papillion",
    "sprocker", "spoodle", "cocker", "pointer", "ridgeback", "bichon",
    "cattle", "shih tzu", "pinscher", "jrt", "jack russell", "corgi",
    "leonberger", "cavalier", "sharpei", "shar pei", "spitz", "pug",
    "maltese", "kelpie", "border collie", "golden", "groenendael", "sprocker",
    "double doodle", "airedale", "harrier", "field retriever", "mixed breed",
    "cross", "rescue", "spca",
]
BREED_RE = re.compile(r"(" + "|".join(re.escape(w) for w in BREED_WORDS) + r")", re.I)
CROSS_RE = re.compile(r"\bx\b", re.I)

# Values that are payment / vaxx / admin flags, not classifiable data. Kept in notes.
NAME_JUNK_TOKENS = ("volunteer", "intern", "rehomed", "moved to contact",
                    "hoping for", "wants a", "have let", "last class",
                    "to join", "joining for")


def is_datetime(v):
    return isinstance(v, (datetime.datetime, datetime.date))


def fmt_date(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()


def looks_like_date(v):
    if is_datetime(v):
        return True
    if isinstance(v, str) and DATE_STR_RE.match(v):
        return True
    return False


def looks_like_phone(v):
    """NZ phone. Called only AFTER a date test has failed."""
    if is_datetime(v):
        return False
    if isinstance(v, (int, float)):
        # numeric phone (waitlist has these), but not a bare year
        iv = int(v)
        return iv >= 10_000_000
    if not isinstance(v, str):
        return False
    s = v.strip()
    if "@" in s:
        return False
    if not PHONE_RE.match(s):
        return False
    digits = re.sub(r"\D", "", s)
    return 7 <= len(digits) <= 12


EMAIL_RE = re.compile(r"[\w.+'-]+@[\w-]+\.[\w.-]+")


def find_email(v):
    """Return a real email address in the cell, or None.

    Not just "contains @": the sheet is full of notes like "12 weeks @ 05/09"
    (@ meaning "as at") that must NOT be treated as emails.
    """
    if not isinstance(v, str):
        return None
    m = EMAIL_RE.search(v)
    return m.group(0) if m else None


def looks_like_email(v):
    return find_email(v) is not None


def looks_like_age(v):
    if not isinstance(v, str):
        return False
    if "@" in v:
        return False
    return bool(AGE_RE.search(v)) and (bool(re.search(r"\d", v)) or
                                       "unknown" in v.lower())


def looks_like_address(v):
    if not isinstance(v, str):
        return False
    s = v.strip()
    if "@" in s or len(s) < 5:
        return False
    low = s.lower()
    if STREET_RE.search(s) and re.search(r"\d", s):
        return True
    if any(sub in low for sub in SUBURBS) and len(s) > 6:
        return True
    return False


def looks_like_breed(v):
    if not isinstance(v, str):
        return False
    s = v.strip()
    if "@" in s:
        return False
    if BREED_RE.search(s):
        return True
    # bare cross like "Lab X" / "Foxy X Staffy"
    if CROSS_RE.search(s) and len(s) <= 40 and re.search(r"[A-Za-z]", s):
        return True
    return False


DOG_STOP = {
    "tbc", "?", "volunteer", "unknown", "yes", "no", "y", "n", "paid",
    "credit", "na", "none", "nil", "full", "ok", "reminder", "intern",
    "yes + faqs", "yes + faq's",
}
DOG_REJECT_RE = re.compile(
    r"(paid|faq|reminder|itinerary|emailed|welcome|\$|r\.s|to pay|"
    r"switch|missing|miss )", re.I)


def looks_like_dog_name(v):
    """Puppy name: short, alphabetic-ish token, not a sentence/flag."""
    if not isinstance(v, str):
        return False
    s = v.strip()
    if "@" in s or len(s) == 0 or len(s) > 25:
        return False
    low = s.lower()
    if low in DOG_STOP or DOG_REJECT_RE.search(s):
        return False
    # mostly letters / parentheses, at most 3 words
    if len(s.split()) > 3:
        return False
    letters = [c for c in s if c.isalpha()]
    if not letters:
        return False
    return sum(c.isalpha() or c in " ()'/-." for c in s) / len(s) > 0.7


def is_header_label_row(vals):
    a = vals.get("A")
    if not isinstance(a, str):
        return False
    if a.strip().lower() not in ("name", "nname"):
        return False
    # confirm it's the column-label row (has Email/Phone labels somewhere)
    joined = " ".join(str(x).lower() for x in vals.values())
    return "email" in joined or "phone" in joined or "puppy" in joined


def is_bare_year(vals):
    if len(vals) != 1 or "A" not in vals:
        return False
    a = vals["A"]
    if isinstance(a, (int, float)) and 2000 <= int(a) <= 2100:
        return True
    if isinstance(a, str) and re.fullmatch(r"20\d\d", a.strip()):
        return True
    return False


def is_course_header(a):
    if not isinstance(a, str):
        return False
    s = a.strip()
    if s.lower().startswith("starts "):
        return True
    # bare weekday header must also carry a time and a month to avoid catching
    # note sentences like "Last class to run ... on Saturday 4th January at 9am"
    if HEADER_RE.match(s) and TIME_RE.search(s) and re.search(MONTH, s, re.I):
        # must START with a weekday (not "Last class ...")
        first = s.split()[0].lower()
        if re.fullmatch(WEEKDAY, first):
            return True
    return False


# ---------------------------------------------------------------------------
# Course header parsing
# ---------------------------------------------------------------------------

def parse_header(a):
    """Extract day / date / time / nickname from a course header string."""
    s = a.strip()
    raw = s
    # nickname: text inside quotes (straight or curly), single or double
    nick = None
    m = re.search(r"['\"‘’“”]([^'\"‘’“”]+)"
                  r"['\"‘’“”]", s)
    if m:
        nick = m.group(1).strip()
    else:
        # sometimes " - The Xyz's" with no quotes
        m2 = re.search(r"-\s*(The [A-Z][\w' ]+)$", s)
        if m2:
            nick = m2.group(1).strip()
    day = None
    md = re.search(WEEKDAY, s, re.I)
    if md:
        day = md.group(1).capitalize()
    # date: weekday + ordinal day + month
    date = None
    mdate = re.search(r"(\d{1,2}(?:st|nd|rd|th)?\s+" + MONTH + r")", s, re.I)
    if mdate:
        date = mdate.group(1).strip()
    # time: first time-range
    time = None
    mt = re.search(r"(\d{1,2}[:.]?\d{0,2}\s*(?:am|pm)?\s*[-–]\s*"
                   r"\d{1,2}[:.]?\d{0,2}\s*(?:am|pm))", s, re.I)
    if mt:
        time = mt.group(1).strip()
    else:
        mt2 = TIME_RE.search(s)
        if mt2:
            time = mt2.group(0).strip()
    return day, date, time, nick, raw


VENUE_HINTS = ("bethlehem", "matua", "bch", "hall")


def header_location(gval):
    """G on an Early Learning header: a venue vs a note."""
    if not isinstance(gval, str):
        return None
    low = gval.lower()
    if any(h in low for h in VENUE_HINTS) and len(gval) < 30 \
            and "cancel" not in low and "sent" not in low \
            and "welcome" not in low and "week" not in low.split("+")[0][:0] \
            and "itinerary" not in low:
        return gval.strip()
    return None


# ---------------------------------------------------------------------------
# Enrolment classification
# ---------------------------------------------------------------------------

def classify_row(vals):
    """Classify a data row's cells into an enrolment dict."""
    rec = {
        "name": None, "email": None, "phone": None, "address": None,
        "dog": None, "breed": None, "age": None, "dateBooked": None,
        "notes": None,
    }
    notes = []

    # A is always the name
    a = vals.get("A")
    rec["name"] = (a.strip() if isinstance(a, str) else a)

    breed_taken = False
    dog_taken = False

    for col in sorted(vals.keys(), key=lambda c: openpyxl.utils.column_index_from_string(c)):
        if col == "A":
            continue
        v = vals[col]
        if v is None:
            continue
        # Order matters: email -> date -> phone -> age -> breed -> address -> dog
        if rec["email"] is None:
            em = find_email(v)
            if em:
                rec["email"] = em
                # if the cell held extra text beyond the address, keep it
                extra = v.replace(em, "").strip(" /,;|") if isinstance(v, str) else ""
                if extra and len(extra) > 2:
                    notes.append(f"{col}:{extra}")
                continue
        if rec["dateBooked"] is None and looks_like_date(v):
            rec["dateBooked"] = fmt_date(v)
            continue
        if rec["phone"] is None and looks_like_phone(v):
            rec["phone"] = (str(int(v)) if isinstance(v, (int, float))
                            else v.strip())
            continue
        if rec["age"] is None and looks_like_age(v):
            rec["age"] = v.strip()
            continue
        if not breed_taken and rec["breed"] is None and looks_like_breed(v):
            rec["breed"] = v.strip()
            breed_taken = True
            continue
        if rec["address"] is None and looks_like_address(v):
            rec["address"] = v.strip()
            continue
        if not dog_taken and rec["dog"] is None and looks_like_dog_name(v):
            rec["dog"] = v.strip()
            dog_taken = True
            continue
        # leftover -> note, tagged with its column
        sval = fmt_date(v) if is_datetime(v) else str(v).strip()
        if sval:
            notes.append(f"{col}:{sval}")

    rec["notes"] = " | ".join(notes) if notes else None
    return rec


def compact_raw(vals):
    parts = []
    for col in sorted(vals.keys(), key=lambda c: openpyxl.utils.column_index_from_string(c)):
        v = vals[col]
        sval = fmt_date(v) if is_datetime(v) else str(v).strip()
        parts.append(f"{col}={sval}")
    return " | ".join(parts)


# ---------------------------------------------------------------------------
# Main extraction
# ---------------------------------------------------------------------------

def norm_name(n):
    if not n:
        return ""
    return re.sub(r"\s+", " ", str(n).strip().lower())


def norm_phone(p):
    if not p:
        return ""
    return re.sub(r"\D", "", str(p))


COUPLE_SEP_RE = re.compile(r"[A-Za-z]\s*(?:&|/|\+| and )\s*[A-Za-z]")


def name_has_multiple(name):
    """A single key covering multiple people (a couple / household)."""
    if not isinstance(name, str):
        return False
    return bool(COUPLE_SEP_RE.search(name))


def main():
    wb = openpyxl.load_workbook(SOURCE, data_only=True)

    courses = []
    enrolments = []
    needs_review = []
    waitlist = []
    course_id = 0

    # ---- class sheets -----------------------------------------------------
    for sheet in CLASS_SHEETS:
        ws = wb[sheet]
        cur_course = None
        for r in range(1, ws.max_row + 1):
            vals = {}
            for c in range(1, MAX_COL):
                v = ws.cell(row=r, column=c).value
                if v is not None and not (isinstance(v, str) and v.strip() == ""):
                    vals[get_column_letter(c)] = v
            if not vals:
                continue

            a = vals.get("A")

            # course header?
            if is_course_header(a):
                course_id += 1
                day, date, time, nick, raw = parse_header(a)
                loc = None
                if sheet == "Early Learning":
                    loc = header_location(vals.get("G"))
                cur_course = course_id
                courses.append(OrderedDict([
                    ("id", course_id),
                    ("sheet", sheet),
                    ("nickname", nick),
                    ("day", day),
                    ("date", date),
                    ("time", time),
                    ("location", loc),
                    ("headerRaw", raw),
                ]))
                continue

            # skip column-label rows and bare years
            if is_header_label_row(vals) or is_bare_year(vals):
                continue

            # only rows inside a course become enrolments
            if cur_course is None:
                continue

            rec = classify_row(vals)
            src_code = f"{SHEET_PREFIX[sheet]}{r}"

            if rec["email"] is None and rec["phone"] is None:
                needs_review.append(OrderedDict([
                    ("sheet", sheet),
                    ("row", r),
                    ("reason", "no email+phone"),
                    ("raw", compact_raw(vals)),
                ]))
                continue

            enrolments.append(OrderedDict([
                ("courseId", cur_course),
                ("name", rec["name"]),
                ("email", rec["email"]),
                ("phone", rec["phone"]),
                ("address", rec["address"]),
                ("dog", rec["dog"]),
                ("breed", rec["breed"]),
                ("age", rec["age"]),
                ("dateBooked", rec["dateBooked"]),
                ("notes", rec["notes"]),
                ("sourceSheet", sheet),
                ("sourceRow", r),
                ("sourceCode", src_code),
            ]))

    # ---- waitlist ---------------------------------------------------------
    ws = wb["Waitlist"]
    category = None
    last_lead = None
    for r in range(1, ws.max_row + 1):
        vals = {}
        for c in range(1, MAX_COL):
            v = ws.cell(row=r, column=c).value
            if v is not None and not (isinstance(v, str) and v.strip() == ""):
                vals[get_column_letter(c)] = v
        if not vals:
            continue
        a = vals.get("A")

        # column-label row
        if is_header_label_row(vals):
            continue

        # category header: A all-caps, <=2 cells
        if isinstance(a, str):
            letters = [ch for ch in a if ch.isalpha()]
            upfrac = (sum(ch.isupper() for ch in letters) / len(letters)
                      if letters else 0)
            if len(vals) <= 2 and upfrac > 0.7 and len(a) < 40 and letters:
                category = a.strip()
                last_lead = None
                continue

        if category is None:
            continue

        # continuation / note-only row (no name, no contact) -> attach to prev
        rec = classify_row(vals)
        aname = rec["name"] if isinstance(rec["name"], str) else ""
        low = aname.lower()
        is_note_sentence = (
            aname and rec["email"] is None and rec["phone"] is None and
            (any(t in low for t in ("hoping", "wants a", "have let",
                                    "let them", "let her", "let him"))
             or len(aname.split()) > 4)
        )
        if is_note_sentence and last_lead is not None:
            extra = compact_raw(vals)
            last_lead["notes"] = ((last_lead["notes"] + " || " + extra)
                                  if last_lead["notes"] else extra)
            continue

        # a real lead needs a name that looks like a person
        if not aname:
            continue

        lead = OrderedDict([
            ("category", category),
            ("name", aname),
            ("phone", rec["phone"]),
            ("email", rec["email"]),
            ("address", rec["address"]),
            ("dog", rec["dog"]),
            ("breedAge", rec["breed"] or rec["age"]),
            ("notes", rec["notes"]),
            ("sourceRow", r),
        ])
        waitlist.append(lead)
        last_lead = lead

    # ---- dedup clients ----------------------------------------------------
    clients = []
    by_email = {}
    by_phone = {}
    by_name = {}

    def get_client(rec):
        em = norm_name(rec["email"]) if rec["email"] else ""
        ph = norm_phone(rec["phone"]) if rec["phone"] else ""
        nm = norm_name(rec["name"])
        # email is the strongest key
        if em and em in by_email:
            return by_email[em]
        if ph and ph in by_phone:
            return by_phone[ph]
        if not em and not ph and nm and nm in by_name:
            return by_name[nm]
        # new client
        cl = OrderedDict([
            ("name", rec["name"]),
            ("email", rec["email"]),
            ("phone", rec["phone"]),
            ("address", rec["address"]),
            ("dogs", []),
            ("courseIds", []),
            ("notes", ""),
            ("sourceRows", []),
            ("multipleNamesOnKey", False),
            ("_names", set()),
            ("_notes", []),
        ])
        clients.append(cl)
        if em:
            by_email[em] = cl
        if ph:
            by_phone[ph] = cl
        if nm:
            by_name.setdefault(nm, cl)
        return cl

    for e in enrolments:
        cl = get_client(e)
        # fill blanks
        if not cl["email"] and e["email"]:
            cl["email"] = e["email"]
            by_email.setdefault(norm_name(e["email"]), cl)
        if not cl["phone"] and e["phone"]:
            cl["phone"] = e["phone"]
            by_phone.setdefault(norm_phone(e["phone"]), cl)
        if not cl["address"] and e["address"]:
            cl["address"] = e["address"]
        if e["dog"] and e["dog"] not in cl["dogs"]:
            cl["dogs"].append(e["dog"])
        if e["courseId"] not in cl["courseIds"]:
            cl["courseIds"].append(e["courseId"])
        cl["sourceRows"].append(e["sourceCode"])
        cl["_names"].add(e["name"])
        if e["notes"]:
            cl["_notes"].append(f"[{e['sourceCode']}] {e['notes']}")
        if name_has_multiple(e["name"]):
            cl["multipleNamesOnKey"] = True

    # finalize clients
    for cl in clients:
        # a couple/household = any source name for this key carries a separator
        # (two people on one email/phone). NOT spelling variants of one person.
        if any(name_has_multiple(n) for n in cl["_names"]):
            cl["multipleNamesOnKey"] = True
        cl["notes"] = " || ".join(dict.fromkeys(cl["_notes"]))
        del cl["_names"]
        del cl["_notes"]

    result = OrderedDict([
        ("courses", courses),
        ("enrolments", enrolments),
        ("clients", clients),
        ("waitlist", waitlist),
        ("needsReview", needs_review),
    ])

    with open(JSON_OUT, "w") as f:
        json.dump(result, f, indent=2, ensure_ascii=False)

    build_workbook(result)

    # ---- summary ----------------------------------------------------------
    multi = sum(1 for c in clients if c["multipleNamesOnKey"])
    repeat = sum(1 for c in clients if len(c["courseIds"]) > 1)
    with_email = sum(1 for e in enrolments if e["email"])
    print("=== PRIME PUPS extraction summary ===")
    print(f"courses           : {len(courses)}")
    print(f"enrolment rows     : {len(enrolments)}")
    print(f"  with real email  : {with_email}")
    print(f"unique clients     : {len(clients)}")
    print(f"  did >1 course    : {repeat}")
    print(f"  multiNamesOnKey  : {multi}")
    print(f"waitlist leads     : {len(waitlist)}")
    print(f"needs-review rows  : {len(needs_review)}")
    print(f"\nJSON  -> {JSON_OUT}")
    print(f"XLSX  -> {XLSX_OUT}")
    return result


# ---------------------------------------------------------------------------
# Review workbook
# ---------------------------------------------------------------------------

HEAD_FILL = PatternFill("solid", fgColor="1F6F6B")
HEAD_FONT = Font(bold=True, color="FFFFFF")
REVIEW_FILL = PatternFill("solid", fgColor="FFF2CC")


def _style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"


def build_workbook(result):
    wb = openpyxl.Workbook()
    courses_by_id = {c["id"]: c for c in result["courses"]}

    def course_label(cid):
        c = courses_by_id.get(cid)
        if not c:
            return f"#{cid}"
        bits = [f"#{cid}"]
        if c["nickname"]:
            bits.append(c["nickname"])
        if c["date"]:
            bits.append(c["date"])
        bits.append(f"({c['sheet']})")
        return " ".join(bits)

    # --- Clients (deduped) ---
    ws = wb.active
    ws.title = "Clients (deduped)"
    cols = ["Name", "Email", "Phone", "Address", "Dog(s)", "#Courses",
            "Courses (history)", "Notes (merged)", "Source rows", "REVIEW?"]
    ws.append(cols)
    for cl in sorted(result["clients"], key=lambda x: norm_name(x["name"])):
        flags = []
        if cl["multipleNamesOnKey"]:
            flags.append("couple/multi-name")
        if not cl["name"]:
            flags.append("no name")
        review = ", ".join(flags)
        row = [
            cl["name"], cl["email"], cl["phone"], cl["address"],
            ", ".join(cl["dogs"]), len(cl["courseIds"]),
            " ; ".join(course_label(i) for i in cl["courseIds"]),
            cl["notes"], ", ".join(cl["sourceRows"]), review,
        ]
        ws.append(row)
        if review:
            for c in range(1, len(cols) + 1):
                ws.cell(row=ws.max_row, column=c).fill = REVIEW_FILL
    _style_header(ws, len(cols))
    widths = [22, 30, 16, 34, 18, 9, 40, 50, 20, 9]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # --- Courses ---
    ws = wb.create_sheet("Courses")
    cols = ["ID", "Sheet", "Nickname", "Day", "Date", "Time", "Location",
            "#Enrolled", "Header (raw)"]
    ws.append(cols)
    enrol_count = {}
    for e in result["enrolments"]:
        enrol_count[e["courseId"]] = enrol_count.get(e["courseId"], 0) + 1
    for c in result["courses"]:
        ws.append([c["id"], c["sheet"], c["nickname"], c["day"], c["date"],
                   c["time"], c["location"], enrol_count.get(c["id"], 0),
                   c["headerRaw"]])
    _style_header(ws, len(cols))
    for i, w in enumerate([5, 16, 22, 10, 22, 20, 16, 10, 50], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # --- Enrolments ---
    ws = wb.create_sheet("Enrolments")
    cols = ["CourseId", "Course label", "Client name", "Email", "Phone",
            "Dog", "Breed", "Age", "Date booked", "Notes", "Source"]
    ws.append(cols)
    for e in result["enrolments"]:
        ws.append([e["courseId"], course_label(e["courseId"]), e["name"],
                   e["email"], e["phone"], e["dog"], e["breed"], e["age"],
                   e["dateBooked"], e["notes"], e["sourceCode"]])
    _style_header(ws, len(cols))
    for i, w in enumerate([9, 34, 22, 30, 16, 14, 22, 14, 14, 40, 8], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # --- Waitlist ---
    ws = wb.create_sheet("Waitlist (leads)")
    cols = ["Category", "Name", "Phone", "Email", "Address", "Dog",
            "Breed/Age", "Notes", "Source row"]
    ws.append(cols)
    for w_ in result["waitlist"]:
        ws.append([w_["category"], w_["name"], w_["phone"], w_["email"],
                   w_["address"], w_["dog"], w_["breedAge"], w_["notes"],
                   w_["sourceRow"]])
    _style_header(ws, len(cols))
    for i, w in enumerate([16, 22, 16, 30, 34, 14, 24, 44, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # --- Needs-Review ---
    ws = wb.create_sheet("Needs-Review")
    cols = ["Sheet", "Row", "Reason", "Raw"]
    ws.append(cols)
    for n in result["needsReview"]:
        ws.append([n["sheet"], n["row"], n["reason"], n["raw"]])
    _style_header(ws, len(cols))
    for i, w in enumerate([16, 8, 20, 90], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(XLSX_OUT)


if __name__ == "__main__":
    main()
